"""
test_pivot_extract.py
---------------------
Regression suite. Each fixture is a real .xlsx and each test asserts the GATE
VERDICT, not just that parsing "worked" -- a file that parses cleanly and loads
wrong numbers is the failure this module exists to prevent, so "did it block?"
is the only assertion worth making.

Fixture families:
  originals  -- the four SME samples (1, 2, 3, 6)
  round 1    -- decode mutations (indent, blanked repeats, negatives, naming)
  round 2    -- shape/identity mutations (compact, numeric dims, subtotals,
                deep nesting, dirty member values)

THE HEADLINE ASSERTION is test_dirty_member_values_are_blocked: that file parses
perfectly, ties to the source total exactly, has no nulls -- and turns three
departments into six. It used to PASS. If it ever passes again, the identity gate
has regressed and the warehouse is taking corrupted keys.

Run:  pytest -v test_pivot_extract.py
"""

import pandas as pd
import pytest

import pivot_extract as px


CONFIG = "configuration.json"
UPLOADS = "/mnt/user-data/uploads"
SOURCE_TOTAL_FY26 = 86273          # the pivot's own declared grand total


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def cfg():
    """The SME declaration. Loading it also validates every normalization
    license against its own vocabulary -- a contradictory spec fails here."""
    return px.load_configuration(CONFIG)


def run(path, cfg):
    """Parse one file and return (tidy, report, high_rows)."""
    tidy, report = px.parse(
        path, config=cfg, var_name="FiscalYear", value_name="Value",
        return_report=True,
    )
    return tidy, report, px.review_items(report)


def leaf_sum(tidy, period="FY26"):
    """Sum ONLY the leaves -- parents are subtotals of their children."""
    fy = tidy[tidy.FiscalYear == period]
    return fy[fy.is_leaf].Value.sum()


def methods(high):
    return set(high.method)


# ---------------------------------------------------------------------------
# Configuration: the declaration must be self-consistent before any file is read
# ---------------------------------------------------------------------------

def test_config_loads_and_licenses_are_self_consistent(cfg):
    assert cfg["hierarchy_levels"] == ["DEPT", "Service", "Variant"]
    assert cfg["vocabularies"]["Service"] == {"MC", "N", "F", "F1", "F2", "F3"}
    # Service licenses punctuation-stripping; that is only legal because no two
    # declared Service members collide under it.
    ok, collisions = px.validate_license(
        cfg["vocabularies"]["Service"], cfg["licenses"]["Service"]
    )
    assert ok and not collisions


def test_license_that_would_merge_two_members_is_rejected():
    """The machine does not decide what is ignorable -- it decides whether YOUR
    answer is coherent. Declaring F-1 and F1 as distinct members while also
    licensing punctuation-stripping is a contradiction, and must not start."""
    ok, collisions = px.validate_license(
        {"F-1", "F1"}, {"punctuation": True, "case": False}
    )
    assert not ok
    assert collisions


# ---------------------------------------------------------------------------
# Value healing: heal what is unambiguous, decide what is a decision
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("raw,expected", [
    ("A ", "A"),                 # trailing space
    ("\u00a0C", "C"),            # non-breaking space
    ("A\u200b", "A"),            # zero-width space: invisible, breaks equality
    ("A\nB", "A B"),             # wrapped-cell line break
])
def test_universal_healing_is_information_preserving(raw, expected):
    assert px.heal_universal(raw) == expected


def test_licensed_healing_is_off_unless_declared():
    assert px.heal_licensed("F-1", {"punctuation": False}) == "F-1"
    assert px.heal_licensed("F-1", {"punctuation": True}) == "F1"


def test_accounting_negative_is_a_number_not_a_hole():
    """(948) is negative in every finance export ever written. Reading it as
    unparseable is not conservative, it is wrong."""
    assert px.to_number("(948)") == -948.0
    assert px.to_number("$1,234.50") == 1234.5
    assert px.to_number(".5") == 0.5
    assert pd.isna(px.to_number("n/a"))


# ---------------------------------------------------------------------------
# Blank semantics: suppressed repeat vs. genuinely missing
# ---------------------------------------------------------------------------

def test_left_prefix_blank_is_filled_at_the_same_level():
    """(blank, N, C) is a suppressed repeat. The fill must come from the last
    label at the SAME indent level -- a naive ffill grabs the indented CHILD
    above and writes 'A subtype 2' into A's siblings, while every sum still
    ties. Silent-wrong, in exactly the way this module exists to prevent."""
    body = pd.DataFrame({
        "DEPT":    ["A", "A subtype 1", None, None],
        "Service": ["MC", "MC", "N", "F"],
    })
    out, recs = px.resolve_blank_dimensions(
        body, ["DEPT", "Service"], levels=[0, 1, 0, 0]
    )
    assert list(out.DEPT) == ["A", "A subtype 1", "A", "A"]
    assert any(r["method"] == px.LABEL_REPEAT_FILLED for r in recs)


def test_blank_with_populated_cell_to_its_left_is_never_filled():
    """(A, blank, C) cannot arise from suppression -- Excel only blanks a label
    when the entire prefix above it is unchanged. So this is a real hole, and
    filling it would fabricate data."""
    body = pd.DataFrame({"DEPT": ["A"], "Service": [None], "Variant": ["C"]})
    out, recs = px.resolve_blank_dimensions(body, ["DEPT", "Service", "Variant"])
    assert pd.isna(out.Service.iloc[0])                      # NOT invented
    assert any(r["method"] == px.DIM_MISSING_VALUE for r in recs)


# ---------------------------------------------------------------------------
# The four originals
# ---------------------------------------------------------------------------

def test_sample_6_is_clean(cfg):
    tidy, _report, high = run(f"{UPLOADS}/Extraction_Sample_6.xlsx", cfg)
    assert len(high) == 0, f"unexpected blocks: {methods(high)}"
    assert leaf_sum(tidy) == SOURCE_TOTAL_FY26


@pytest.mark.parametrize("name", [
    "Extraction_Sample_1.xlsx",
    "Extraction_Sample_2.xlsx",
    "Extraction_Sample_3.xlsx",
])
def test_legacy_samples_block_only_on_the_missing_metablock(cfg, name):
    """The pre-metablock files. They parse and reconcile perfectly; they block
    solely because the metadata contract is not satisfied. Expected."""
    tidy, _report, high = run(f"{UPLOADS}/{name}", cfg)
    assert methods(high) == {"metadata_missing"}
    assert leaf_sum(tidy) == SOURCE_TOTAL_FY26


# ---------------------------------------------------------------------------
# Round 1 -- decode mutations
# ---------------------------------------------------------------------------

def test_real_excel_indent_attribute_parses(cfg):
    """The SME templates fake indentation with leading spaces, so the native
    Excel indent property was never exercised in dev. It works."""
    tidy, _r, high = run("m1_real_excel_indent.xlsx", cfg)
    assert len(high) == 0
    assert leaf_sum(tidy) == SOURCE_TOTAL_FY26


def test_blanked_repeat_labels_fill_correctly(cfg):
    """Excel's DEFAULT. 'Repeat All Item Labels' is off unless someone ticks it,
    so the top level arrives blank on 15 of 20 rows. This used to null the DEPT
    of three quarters of the file while every gate passed."""
    tidy, report, high = run("m2_blanked_repeat_labels.xlsx", cfg)
    assert len(high) == 0
    assert tidy.DEPT.isna().sum() == 0
    assert set(tidy[tidy.sub_level == 0].DEPT) == {"A", "B", "C"}   # not 'A subtype 2'
    assert leaf_sum(tidy) == SOURCE_TOTAL_FY26
    assert (report.method == px.LABEL_REPEAT_FILLED).any()          # heal was logged


def test_accounting_negatives_are_read_not_dropped(cfg):
    """Previously '(948)' coerced to NaN, the row vanished, and reconciliation
    reported a mystery delta. Now it reads as -948 and ties to the source."""
    tidy, _r, high = run("m3_paren_negatives.xlsx", cfg)
    assert len(high) == 0, f"unexpected blocks: {methods(high)}"
    assert leaf_sum(tidy) == 84377          # the file's own adjusted total
    assert (tidy.Value < 0).any()


def test_department_named_total_is_not_deleted(cfg):
    """TOTAL_RE used to scan every cell in the row, so a department called
    'Total Rewards' had six rows silently deleted. Now the total test is anchored
    to the outline column: the rows survive, the arithmetic ties, and the gate
    blocks for the RIGHT reason -- the member is not declared."""
    tidy, _r, high = run("m4_dept_named_Total.xlsx", cfg)
    assert leaf_sum(tidy) == SOURCE_TOTAL_FY26        # nothing was deleted
    assert methods(high) == {px.MEMBER_UNDECLARED}    # named, not a mystery delta


def test_numbers_stored_as_text_parse(cfg):
    tidy, _r, high = run("m5_numbers_as_text.xlsx", cfg)
    assert len(high) == 0
    assert leaf_sum(tidy) == SOURCE_TOTAL_FY26


# ---------------------------------------------------------------------------
# Round 2 -- shape and identity mutations
# ---------------------------------------------------------------------------

def test_compact_form_pivot_is_blocked(cfg):
    """All levels collapsed into one indented column: Service and Variant do not
    exist as columns at all."""
    _tidy, _r, high = run("mut/M1_compact_form.xlsx", cfg)
    assert "hierarchy_missing" in methods(high)


def test_numeric_coded_dimensions_are_blocked(cfg):
    """Service=10/20/30 gets classified as a VALUE column and melted into the
    fact. Two independent gates catch it -- the point of defence in depth."""
    _tidy, _r, high = run("mut/M2_numeric_dims.xlsx", cfg)
    assert {"hierarchy_missing", px.DIMENSION_NULL} <= methods(high)


def test_subtotal_rows_are_dropped_and_grand_total_is_the_baseline(cfg):
    """Per-DEPT subtotals plus a Grand Total. Subtotals must be dropped and the
    LAST total row used as the reconciliation baseline."""
    tidy, _r, high = run("mut/M3_subtotal_rows.xlsx", cfg)
    assert len(high) == 0
    assert leaf_sum(tidy) == SOURCE_TOTAL_FY26
    assert len(tidy[tidy.FiscalYear == "FY26"]) == 18       # 18 data rows, no subtotals


def test_deep_nesting_and_blank_literal(cfg):
    """Three indent levels. Ancestry must be exact and intermediate parents must
    be excluded from the leaf sum. The undeclared 'sub-sub' members correctly
    block: a new member is a business decision, not a cell's say-so."""
    tidy, report, high = run("mut/M4_deep_nest_blank.xlsx", cfg)
    assert leaf_sum(tidy) == SOURCE_TOTAL_FY26              # parents not double-counted
    assert (tidy.sub_level == 2).any()                      # depth reached
    assert methods(high) == {px.MEMBER_UNDECLARED}
    # Excel renders a null grouping key as the literal "(blank)"; it becomes a
    # declared sentinel rather than a member named "(blank)".
    assert (report.method == px.MEMBER_BLANK_LITERAL).any()
    assert px.BLANK_MEMBER_SENTINEL in set(tidy.Service)


def test_dirty_member_values_are_blocked(cfg):
    """*** THE HEADLINE REGRESSION ***

    Three departments in, six out: "A " / "A", "b" / "B", "C\\xa0" / "C". The
    shape is perfect. The arithmetic ties to the source total EXACTLY, because
    the sums are conserved -- the money is just attributed to phantom members.
    There are no nulls. The rows are visually identical in any printed report.

    The arithmetic gate cannot see this. The hierarchy gate cannot see this. The
    dimension gate cannot see this. This file PASSED every check before the
    identity gate existed. If it ever passes again, the warehouse is taking
    corrupted keys.
    """
    tidy, report, high = run("mut/M5_dirty_dim_values.xlsx", cfg)

    assert leaf_sum(tidy) == SOURCE_TOTAL_FY26              # arithmetic still ties!
    assert tidy.DEPT.isna().sum() == 0                      # no nulls either!
    assert px.MEMBER_UNDECLARED in methods(high)            # ...and yet: BLOCKED

    # Whitespace / unicode noise heals silently and auditably...
    healed = report[report.method == px.MEMBER_NORMALIZED]
    assert len(healed) >= 2

    # ...but case is a DECISION, not a defect, so it goes to a human -- with the
    # answer handed to them rather than applied on their behalf.
    undeclared = report[report.method == px.MEMBER_UNDECLARED]
    assert "b" in set(undeclared.source_header)
    assert any("'B'" in str(n) for n in undeclared.note)    # the hint names it


# ---------------------------------------------------------------------------
# Fail-closed pipeline + provenance
# ---------------------------------------------------------------------------

def test_ingest_gate_is_on_by_default(cfg):
    """A safety property you have to remember to invoke is not a safety
    property. A batch containing a bad file must RAISE, not return."""
    with pytest.raises(ValueError, match="Load blocked"):
        px.ingest(["mut/M5_dirty_dim_values.xlsx"], config=cfg,
                  var_name="FiscalYear", value_name="Value")


def test_bad_file_does_not_stop_the_batch(cfg):
    """One unreadable file must be reported, not fatal -- a thirty-file batch
    always completes and hands back the list to look at."""
    _data, report = px.ingest(
        [f"{UPLOADS}/Extraction_Sample_6.xlsx", "does_not_exist.xlsx"],
        config=cfg, var_name="FiscalYear", value_name="Value", gate=False,
    )
    assert "file_error" in set(report.method)


def test_reload_is_idempotent(cfg):
    """Re-running a batch must not double the facts. Same bytes -> same hash ->
    same row_keys, so the warehouse MERGEs onto itself instead of appending a
    second copy. This is the silent-wrong that DEFEATS the gates rather than
    tripping them: each file reconciles perfectly on its own."""
    files = [f"{UPLOADS}/Extraction_Sample_6.xlsx"]
    a, _ = px.ingest(files, config=cfg, var_name="FiscalYear", value_name="Value")
    b, _ = px.ingest(files, config=cfg, var_name="FiscalYear", value_name="Value")

    assert a.source_sha256.iloc[0] == b.source_sha256.iloc[0]   # content-addressed
    assert list(a.row_key) == list(b.row_key)                   # stable identity
    assert a.row_key.nunique() == len(a)                        # unique per fact


# ---------------------------------------------------------------------------
# Governance modes: closed / registry / observed
# ---------------------------------------------------------------------------
#
# The mode is chosen by CARDINALITY, and the failure of picking wrong is not
# symmetric:
#   closed on a granular level -> a hand-maintenance treadmill, and a treadmill
#       is not a control because it gets abandoned.
#   observed on a small level  -> no membership check on the dimension everybody
#       actually looks at.

import copy
import json
import os


def _cfg_with(tmp_path, mode, bootstrap=False, level="Service"):
    """Rewrite the declaration with `level` under a different governance mode."""
    c = json.load(open(CONFIG))
    for h in c["hierarchy"]:
        if h["canonical"] == level:
            h["mode"] = mode
            if mode != px.MODE_CLOSED:
                h.pop("vocabulary", None)
            if bootstrap:
                h["bootstrap"] = True
    p = tmp_path / "cfg.json"
    p.write_text(json.dumps(c))
    return px.load_configuration(str(p))


def test_mode_is_explicit_and_validated(tmp_path):
    """'vocabulary is null' quietly meaning 'ungoverned' is exactly the implicit
    rule that leaves a level unchecked by accident. The mode must be stated."""
    c = json.load(open(CONFIG))
    assert all(h["mode"] in px.VALID_MODES for h in c["hierarchy"])

    # closed with no members would reject every file
    c["hierarchy"][0].pop("vocabulary")
    p = tmp_path / "bad.json"
    p.write_text(json.dumps(c))
    with pytest.raises(ValueError, match="no vocabulary"):
        px.load_configuration(str(p))


def test_registry_bootstrap_then_blocks_new_members(tmp_path):
    """First load of a granular level: EVERYTHING is new. You cannot approve five
    thousand codes one at a time, and a reviewer facing five thousand HIGH rows
    turns the gate off. So bootstrap registers wholesale, loudly -- then off."""
    reg = {}
    cfg = _cfg_with(tmp_path, px.MODE_REGISTRY, bootstrap=True)
    _t, report = px.parse(f"{UPLOADS}/Extraction_Sample_6.xlsx", config=cfg,
                          registry=reg, return_report=True)
    assert len(px.review_items(report)) == 0                    # bootstrap passes
    assert (report.method == px.MEMBER_REGISTERED).any()        # ...but loudly
    assert px.approved_members(reg, "Service") == {"MC", "N", "F", "F1", "F2", "F3"}

    # Bootstrap off: a member NOT in the register now blocks.
    cfg = _cfg_with(tmp_path, px.MODE_REGISTRY)
    reg["levels"]["Service"]["members"].pop("F3")               # simulate a new code
    _t, report = px.parse(f"{UPLOADS}/Extraction_Sample_6.xlsx", config=cfg,
                          registry=reg, return_report=True)
    high = px.review_items(report)
    assert px.MEMBER_PENDING in set(high.method)
    assert px.pending_members(reg) == {"Service": ["F3"]}       # the worklist


def test_registry_bulk_approval_is_one_decision(tmp_path):
    """Approval must be cheaper than the workaround. If clearing a new member is
    a week-long ticket, someone widens the vocabulary to a wildcard and you are
    back to six departments named 'A'."""
    reg = {}
    boot = _cfg_with(tmp_path, px.MODE_REGISTRY, bootstrap=True)
    px.parse(f"{UPLOADS}/Extraction_Sample_6.xlsx", config=boot, registry=reg,
             return_report=True)
    reg["levels"]["Service"]["members"].pop("F2")
    reg["levels"]["Service"]["members"].pop("F3")

    cfg = _cfg_with(tmp_path, px.MODE_REGISTRY)
    _t, report = px.parse(f"{UPLOADS}/Extraction_Sample_6.xlsx", config=cfg,
                          registry=reg, return_report=True)
    assert len(px.pending_members(reg)["Service"]) == 2

    px.approve_members(reg, "Service", approver="jane.smith")   # ONE call, all of them
    assert px.pending_members(reg) == {}

    _t, report = px.parse(f"{UPLOADS}/Extraction_Sample_6.xlsx", config=cfg,
                          registry=reg, return_report=True)
    assert len(px.review_items(report)) == 0                    # never asked again


def test_observed_mode_catches_near_duplicates(tmp_path):
    """The check that MATTERS MORE as cardinality rises, not less. With three
    departments a human eventually notices 'b' and 'B' both exist. With five
    thousand project codes nobody will ever look -- and they will split every
    rollup that touches them, silently, forever."""
    import shutil
    from openpyxl import load_workbook

    src = str(tmp_path / "dirty.xlsx")
    shutil.copy(f"{UPLOADS}/Extraction_Sample_6.xlsx", src)
    wb = load_workbook(src)
    wb["F-DEPT-VAR-TEAM-SPRL"].cell(row=20, column=2, value="f-1")   # vs 'F1'
    wb.save(src)

    cfg = _cfg_with(tmp_path, px.MODE_OBSERVED)
    _t, report = px.parse(src, config=cfg, registry={}, return_report=True)
    high = px.review_items(report)
    assert px.MEMBER_NEAR_DUPLICATE in set(high.method)
    note = " ".join(str(n) for n in high.note)
    assert "F1" in note and "f-1" in note


def test_observed_mode_flags_cardinality_drift(tmp_path):
    """3 departments became 6 -- catchable without knowing what a department is."""
    reg = {"levels": {"Service": {"cardinality": 3}}}
    recs = px._check_cardinality_drift(reg, "Service", 6, 0.25)
    assert recs and recs[0]["method"] == px.CARDINALITY_DRIFT
