"""
extract_pivot.py
----------------
Extract a wide "pivot-style" table from an Excel sheet and normalize it into a
tidy, standard columnar (long) DataFrame.

The extractor is position-independent: it does NOT assume the header sits on a
fixed row or that there is a fixed number of dimension / value columns. It:

  1. Scans rows to locate the real header row (skipping junk metadata at the top).
  2. Splits columns into dimension (id) columns vs. value columns by inspecting
     the data types of the values beneath each header.
  3. Drops total/subtotal rows (e.g. "GRAND TOTAL") and fully-blank rows.
  4. Unpivots the value columns into two columns: one for the former column
     header (e.g. fiscal year) and one for the numeric value.
  5. (Optional) Maps the detected header names onto canonical names you supply,
     using explicit aliases first and a fuzzy fallback for near matches.

Public API:
    extract_pivot(path, sheet=0, value_name="Value", var_name="Period",
                  aliases=None, fuzzy=True,
                  accept_threshold=DEFAULT_ACCEPT_THRESHOLD,
                  review_threshold=DEFAULT_REVIEW_THRESHOLD,
                  strip_prefixes=DEFAULT_AGG_PREFIXES,
                  return_report=False) -> pd.DataFrame
    ingest(paths, ...)         -> (data_df, report_df)   # mass ingest many files
    resolve_headers(headers, ...) -> list[dict]          # per-header report rows
    review_items(report)       -> DataFrame of HIGH-priority rows

================================================================================
Fuzzy header matching — how it works
================================================================================
Real-world exports rarely spell a column exactly the way your schema does. When
an incoming header does not match one of your canonical names or aliases exactly,
the resolver falls back to *fuzzy* matching to catch typos, casing differences,
extra spaces, and small wording changes (e.g. "Divison" for "Division").

  * Normalization first. Both the incoming header and every candidate name are
    lower-cased, have runs of whitespace collapsed to a single space, and (for
    the incoming header) have any leading aggregation prefix like "Sum of"
    stripped. So casing / spacing / prefix differences never count against a
    match — only the meaningful text is compared.

  * Similarity score. Each normalized header is compared against every candidate
    using Python's standard-library difflib.SequenceMatcher, which returns a
    ratio between 0.0 and 1.0. That ratio is 2*M / T, where T is the total number
    of characters in the two strings and M is the number of characters that line
    up in matching runs (the Ratcliff/Obershelp algorithm). 1.0 means identical;
    the closer to 1.0, the more alike. The candidate with the highest ratio wins.
    (This is pure-Python and needs no extra packages — no Levenshtein library.)

  * Two-threshold decision. The winning ratio is compared against two cutoffs,
    which is what drives the kickback report:
      - ratio >= accept_threshold (default 0.85): confident match, applied
        silently and logged as method "fuzzy" with review "OK".
      - review_threshold (default 0.60) <= ratio < accept_threshold: the match is
        still applied so ingestion is not blocked, but the row is flagged "HIGH"
        (method "fuzzy_low_confidence") so a reviewer can confirm it.
      - ratio < review_threshold: no rename happens. The header passes through
        under its cleaned original name; if the closest candidate was at least
        NEARMISS_NOTE_FLOOR (0.40) similar, that near-miss is noted for context.

  * Collisions. If two different source headers end up pointing at the same name,
    the higher-confidence one keeps it and the other is reverted to its original
    name and flagged "HIGH" (method "collision_reverted"), so two columns are
    never silently merged into one.

Tune accept_threshold / review_threshold to make matching stricter or looser, or
pass fuzzy=False to disable fuzzy matching entirely (exact + alias only).
"""

from __future__ import annotations
import re
import os
import json
import functools
import unicodedata
from difflib import SequenceMatcher
import pandas as pd
import numpy as np


# ============================================================================
# GLOBALS — all module-level constants live here, at the top, always.
# ============================================================================

# Regex: matches a "total" / "subtotal" / "grand total" row LABEL (any case).
# NOTE: this is tested against the OUTLINE COLUMN ONLY, never scanned across the
# whole row -- see _is_total_row. Scanning the row deleted any department whose
# name merely contained the word ("Total Rewards"), and let a stray "total" in
# the trailing notes block become the reconciliation baseline.
TOTAL_RE = re.compile(r"\b(grand\s+total|subtotal|total)\b", re.IGNORECASE)

# Regex: a cell whose text is purely a number. Accepts what finance exports
# actually emit:
#   1,234      thousands separators
#   -56.7      leading sign
#   (948)      ACCOUNTING NEGATIVE -- the standard in finance, and previously
#              unmatched: the cell coerced to NaN, the row silently vanished,
#              and reconciliation reported a mystery delta rather than "I cannot
#              read your negatives".
#   $1,234     leading currency symbol
#   .5         bare decimal
NUMERIC_RE = re.compile(
    r"\s*[-+]?[$£€]?\s*(?:\(\s*[$£€]?[\d,]*\.?\d+\s*\)|[\d,]*\.?\d+)\s*"
)


# Converts one cell to a float, honouring accounting negatives. Parentheses mean
# NEGATIVE in every finance export ever written; treating "(948)" as unreadable
# is not a conservative choice, it is a wrong one.
def to_number(x):
    """Coerce a cell to a number. '(948)' -> -948.0. Non-numeric -> NaN."""
    if isinstance(x, (int, float, np.integer, np.floating)) and not isinstance(x, bool):
        return float(x)
    if not isinstance(x, str):
        return pd.NA
    s = x.strip()
    if not NUMERIC_RE.fullmatch(s):
        return pd.NA
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").strip()
    s = re.sub(r"[$£€,\s]", "", s)
    if s.startswith("+"):
        s = s[1:]
    try:
        v = float(s)
    except ValueError:
        return pd.NA
    return -v if neg else v


# Vectorised form of to_number for a whole column.
def to_numeric_column(series):
    """Coerce a column to numbers, honouring accounting negatives."""
    return pd.to_numeric(series.map(to_number), errors="coerce")


# Builds the total-label regex from the DECLARED labels. US-locale strings live
# in configuration.json, not here, so a localized workbook ("Gesamtergebnis") is
# a config edit rather than a code change.
def build_total_re(labels=None):
    """Compile a total/subtotal matcher from the declared label list."""
    labels = labels or ["grand total", "subtotal", "total"]
    alt = "|".join(re.escape(l).replace(r"\ ", r"\s+") for l in labels)
    return re.compile(rf"\b({alt})\b", re.IGNORECASE)


# A row is a TOTAL row when the OUTLINE column carries a total label AND no other
# dimension cell is populated. That is what a real total row looks like; a real
# department named "Total Rewards" always carries a Service and a Variant, so it
# survives. Anchoring here is what stops the parser deleting business rows.
def _is_total_row(row, dim_cols=None, outline=None, total_re=None):
    """True if this row is a total/subtotal row, judged on the outline column."""
    total_re = total_re or TOTAL_RE
    if dim_cols is None or outline is None:      # fallback: legacy whole-row scan
        return any(isinstance(v, str) and total_re.search(v) for v in row.values)
    label = row.get(outline)
    if not (isinstance(label, str) and total_re.search(label)):
        return False
    others = [c for c in dim_cols if c != outline]
    return all(_is_blank(row.get(c)) for c in others)

# Aggregation prefixes that Excel PivotTables prepend to value headers
# (e.g. "Sum of FY26"). Stripped before matching so headers map cleanly.
DEFAULT_AGG_PREFIXES = (
    "sum of", "count of", "average of", "avg of",
    "max of", "min of", "total of", "count distinct of",
)

# Review-priority labels used throughout the kickback report.
REVIEW_OK = "OK"
REVIEW_HIGH = "HIGH"

# Fraction of a column's non-null cells that must be numeric for it to be treated
# as a VALUE column rather than a dimension. This is the knob that decides whether
# a numeric-coded dimension (cost centre, dept number) gets melted into the fact
# by mistake -- so it is declared, not buried.
DEFAULT_VALUE_THRESHOLD = 0.6

# Fuzzy-match decision thresholds (difflib ratio, 0.0 .. 1.0). See the module
# docstring's "Fuzzy header matching" section for the full explanation.
DEFAULT_ACCEPT_THRESHOLD = 0.85   # >= this: confident match, applied silently
DEFAULT_REVIEW_THRESHOLD = 0.60   # >= this but < accept: applied, flagged HIGH
NEARMISS_NOTE_FLOOR = 0.40        # >= this: close enough to note the near-miss

# Column order for the per-column kickback report.
REPORT_COLUMNS = [
    "source_header", "resolved_name", "method", "matched_against",
    "confidence", "review_priority", "note",
]

# The hierarchy metadata columns every harmonized output carries, in order.
# Flat sheets get these too (level 0, leaf, no parent) so a mixed batch always
# has one identical shape.
HIERARCHY_COLUMNS = ["sub_level", "is_parent", "is_leaf", "parent_label", "path"]

# ---- Metadata block (the labelled key/value rows above the pivot) ----------
# Each row is "Label:" in one cell and its value in the cell to its right.
# Labels are matched case/space-insensitively, with or without trailing colon.
REPORT_LABEL = "Report:"
# How many rows above the detected header row to scan for the metadata block.
METADATA_SCAN_ROWS = 12
# Excel's hard cap on sheet-name length; report names must fit so the sheet
# name can mirror the Report: cell exactly.
MAX_SHEET_NAME_LEN = 31

# The metadata fields the contract declares. Every declared field ALWAYS becomes
# a column (null if absent), so the output schema never drifts between files.
# An undeclared label found in the block is a HIGH flag -- that is how a typo
# ("Reportr:") gets caught instead of silently creating a new column.
DEFAULT_METADATA_FIELDS = (
    "Report", "Reporter", "Report Type", "Report Team", "Report Date",
)
# Only these must be present. The rest are optional provenance: absent -> null
# column + a visible (non-blocking) note.
REQUIRED_METADATA_FIELDS = ("Report",)
# Prefix for metadata provenance columns. Keeps them namespaced so a metadata
# label can never collide with a hierarchy/dimension column of the same name
# (e.g. a "Report Type:" field vs. a "Type" pivot column).
METADATA_PREFIX = "meta_"

# ---- Hierarchy-subset override --------------------------------------------
# Value written into a hierarchy column that a report legitimately does not
# carry, when that level is explicitly overridden. Deliberately distinct from a
# naturally-occurring blank so downstream can tell "not tracked by this report"
# from "missing/unknown value".
OVERRIDE_NULL = "NA - Overridden"
# Review status for a level that was missing but explicitly permitted. Not
# silenced -- downgraded from HIGH and still visible in the report.
REVIEW_OVERRIDDEN = "OVERRIDDEN"
# Row-level marker column listing which levels were nulled by override
# (comma-separated; empty string when none).
NULLED_LEVELS_COLUMN = "nulled_levels"

# ---- Reconciliation (the arithmetic gate) ---------------------------------
# Absolute tolerance when comparing leaf sums against the pivot's own total row.
# Non-zero because the data contains .5 values and floats do not compare exactly.
# Too tight -> false failures; too loose -> real errors slip through.
RECON_TOLERANCE = 0.01
# Review status for a period whose leaf sum does not tie to the total row.
# This is the check that catches the dangerous failure: numbers that are wrong
# but plausible. It is HIGH by definition.
RECON_METHOD = "reconciliation_mismatch"

# ---- Value healing (dimension MEMBERS, not headers) ------------------------
# The parser has always normalized HEADERS obsessively and dimension VALUES not
# at all -- but values are what become join keys. "A " and "A" are two different
# departments to a GROUP BY, are visually identical in any report, and preserve
# every sum, so neither the arithmetic gate nor the hierarchy gate can see the
# difference. Healing closes that.
#
# The rule that separates healing from guessing:
#   A normalization is SAFE for a dimension iff applying it to that dimension's
#   declared vocabulary never causes two distinct members to collide.
# Universally-safe healings satisfy this for every possible vocabulary and are
# always on. Conditionally-safe ones must be LICENSED per level and are verified
# against the vocabulary at config load (see validate_license).

# Zero-width characters: invisible, and they make "A" != "A" with no visual cue.
# The nastiest of the lot, because no amount of eyeballing a sheet reveals them.
ZERO_WIDTH_RE = re.compile(r"[\u200b\u200c\u200d\ufeff]")
# Punctuation removed only where a level's license permits it (e.g. "F-1" -> "F1"
# is fine when the vocabulary is {F1, F2, F3}; it is destruction when the
# vocabulary declares F-1 and F1 as distinct members).
PUNCT_RE = re.compile(r"[-_./\\]")
# Conditionally-safe healings, all OFF unless a level's config licenses them.
DEFAULT_LICENSE = {"punctuation": False, "case": False}

# Excel renders a genuinely-null grouping key as the literal text "(blank)".
# That is a SEMANTIC mapping, not a healing, so it is declared here rather than
# smuggled into the normalizer. Mapped to an explicit sentinel so downstream can
# tell "the source had no value" from "the cell was empty because of a repeat".
BLANK_MEMBER_LITERAL = "(blank)"
BLANK_MEMBER_SENTINEL = "NA - Not Provided"

# Report methods for the healing + identity layer.
MEMBER_NORMALIZED = "member_normalized"        # healed to a declared member (OK)
MEMBER_BLANK_LITERAL = "member_blank_literal"  # "(blank)" -> sentinel (OK)
MEMBER_UNDECLARED = "member_undeclared"        # not in the vocabulary (HIGH)
MEMBER_NEAR_DUPLICATE = "member_near_duplicate"  # open domain collision (HIGH)
MEMBER_PENDING = "member_pending_approval"     # new, registry mode (HIGH)
MEMBER_REGISTERED = "member_registered"        # auto-accepted at bootstrap (OK)
MEMBER_PARENT_CHANGE = "member_parent_change"  # member moved parent (HIGH)
CARDINALITY_DRIFT = "cardinality_drift"        # member count jumped (HIGH)

# ---- Governance modes ------------------------------------------------------
# How a level's membership is controlled. The right answer depends entirely on
# CARDINALITY, and getting it wrong in either direction is costly:
#
#   MODE_CLOSED   -- you write the list by hand. Correct for SMALL, STABLE,
#       high-blast-radius levels (Department: 3-20 members, on every dashboard,
#       a corruption is catastrophic). Worth the hand-maintenance precisely
#       because there is so little of it.
#
#   MODE_REGISTRY -- the list BUILDS ITSELF, but a new member needs a signature.
#       Correct for GRANULAR levels (hundreds/thousands of members, churning).
#       First sighting blocks; someone approves in bulk; never asked again. This
#       is the mode that stops a controlled vocabulary becoming a treadmill --
#       and a treadmill is not a control, because it gets abandoned.
#
#   MODE_OBSERVED -- no membership check at all. For levels too granular or too
#       churny to gate on. You still get near-duplicate detection and cardinality
#       drift, which is the ONLY control that works at high cardinality: with
#       three departments a human eventually notices 'b' and 'B' both exist; with
#       five thousand project codes, nobody will ever look.
MODE_CLOSED = "closed"
MODE_REGISTRY = "registry"
MODE_OBSERVED = "observed"
VALID_MODES = (MODE_CLOSED, MODE_REGISTRY, MODE_OBSERVED)

# The member registry: machine-appended, human-approved, never hand-edited. Kept
# OUT of configuration.json on purpose -- config is POLICY (small, stable, owned)
# and the registry is the REGISTER (large, churning). Conflating the two is what
# makes a controlled vocabulary feel unmaintainable.
DEFAULT_REGISTRY_PATH = "members.json"
# A member appearing under a parent it has never been seen under before. Not
# proof of an error -- reorgs happen -- but it is the only signal we have that
# reaches into the misattribution blind spot, and it needs history to work.
DEFAULT_DRIFT_TOLERANCE = 0.25   # +/-25% change in distinct members per level
DIMENSION_NULL = "dimension_null"              # null in a declared level (HIGH)

# ---- Blank semantics in a dimension column ---------------------------------
# A blank dimension cell means one of two completely different things, and the
# difference is structural, so we never have to ask the SME or sniff the file:
#
#   Excel suppresses a repeated label ONLY when the entire prefix above it is
#   unchanged. If DEPT changes, Service always prints. Therefore blanks arising
#   from suppression always form a LEFT-CONTIGUOUS PREFIX of the dimension
#   columns. (blank, N, C) can be a suppressed repeat. (A, blank, C) cannot --
#   that can only be a genuinely missing value.
#
# So: left-prefix blanks are filled (from the last label at the SAME indent
# level -- NOT the cell above, which on an indented sheet is a CHILD row and
# would fill "A subtype 2" into the DEPT of A's siblings). Anything else is a
# real hole and is flagged, never invented.
LABEL_REPEAT_FILLED = "label_repeat_filled"    # suppressed repeat, filled (OK)
DIM_MISSING_VALUE = "dim_missing_value"        # blank that is NOT a repeat (HIGH)

# ---- Provenance (idempotent reloads) ---------------------------------------
# `source_file` alone cannot answer the two questions ETL actually asks:
#   "have I loaded these exact bytes before?"  and  "is this row the same row?"
# Without them, re-running a batch after fixing ONE file re-inserts the other
# twenty-nine. Each file still reconciles perfectly on its own, so no gate fires
# -- the facts just quietly double in the warehouse. That is the most likely
# silent-wrong left in the pipeline, and it defeats the gates rather than
# tripping them, because the error is created downstream of every check.
#
#   source_sha256 : hash of the file BYTES. A resubmitted file with the same name
#                   is visibly a different file. A true re-run is visibly not.
#   ingested_at   : UTC load time -- the audit trail for what passed the gate.
#   row_key       : deterministic hash of (content, dims, period). Lets the
#                   warehouse MERGE instead of INSERT, which is what makes a
#                   reload idempotent rather than additive.
PROVENANCE_COLUMNS = ["source_file", "source_sha256", "ingested_at"]
ROW_KEY_COLUMN = "row_key"

# ---- Configuration ---------------------------------------------------------
# The declaration lives in JSON, owned and versioned by the SMEs. The parser
# obeys it; it does not carry business knowledge of its own.
DEFAULT_CONFIG_PATH = "configuration.json"


# Hashes the file's BYTES, not its name. Two submissions with the same filename
# and different contents are different files, and the warehouse must be able to
# see that.
def file_digest(path) -> str:
    """SHA-256 of the file contents (short form, 16 hex chars)."""
    import hashlib

    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


# Deterministic per-row identity: the same file re-ingested yields the same keys,
# so a reload MERGEs onto itself instead of appending a second copy.
def add_row_keys(df, hierarchy_levels, var_name, value_name):
    """Add a deterministic row_key = hash(source content + dims + period)."""
    import hashlib

    if df.empty:
        return df
    levels = [c for c in (hierarchy_levels or []) if c in df.columns]
    parts = [df.get("source_sha256", pd.Series([""] * len(df), index=df.index))]
    parts += [df[c].astype("string").fillna("~") for c in levels]
    if var_name in df.columns:
        parts.append(df[var_name].astype("string").fillna("~"))
    joined = parts[0].astype("string").fillna("~")
    for p in parts[1:]:
        joined = joined + "|" + p
    df[ROW_KEY_COLUMN] = joined.map(
        lambda s: hashlib.sha256(str(s).encode()).hexdigest()[:16]
    )
    return df


# Resolves the actual worksheet NAME, whether `sheet` was given as an index or a
# name -- needed to enforce the "Report: cell == sheet name" contract.
def _sheet_name_of(path, sheet=0):
    """Return the worksheet's name."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    try:
        if isinstance(sheet, int):
            return wb.sheetnames[sheet]
        return sheet if sheet in wb.sheetnames else None
    finally:
        wb.close()


# ============================================================================
# Metadata block (the "Report:" key/value rows sitting above the pivot)
# ============================================================================

# Turns a metadata label into a safe, namespaced column name.
# "Report Team:" -> "meta_report_team".  The prefix guarantees a metadata field
# can never collide with a pivot/hierarchy column of the same word (e.g. a
# "Report Type:" field vs. the "Type" column that aliases to Variant).
def _meta_column(label, prefix=METADATA_PREFIX):
    """Normalize a metadata label to a column name."""
    base = _norm(label).rstrip(":").strip()
    base = re.sub(r"[^a-z0-9]+", "_", base).strip("_")
    return f"{prefix}{base}"


# Reads the WHOLE metadata block above the pivot: every "Label:" / value pair.
# Located by label, never by fixed cell address, so it survives SMEs adding or
# removing junk rows above. Returns the raw values -- validation happens later.
def read_metadata_block(values, header_row, scan_rows=METADATA_SCAN_ROWS):
    """Scan the rows above `header_row` for "Label:" / value pairs.

    A metadata row is any row where some cell's text ends in ':' and the cell
    to its right holds a value. The whole window above the header is scanned --
    a blank row *inside* the block does not truncate it. The trailing-colon test
    is what separates metadata from junk; any junk row that does look like a
    label gets caught later as `metadata_unexpected`.

    Returns (meta, anchor_row):
      meta       -> {normalized_label: value}, values kept in native type
                    (dates stay datetimes). A label with a blank value maps to "".
      anchor_row -> topmost row index of the block, or None if no block found.
    """
    meta, anchor = {}, None
    start = max(0, header_row - scan_rows)
    for r in range(header_row - 1, start - 1, -1):
        row = values.iloc[r]
        for c in range(len(row) - 1):
            cell = row.iloc[c]
            if not isinstance(cell, str) or not cell.strip().endswith(":"):
                continue
            label = cell.strip().rstrip(":").strip()
            if not label:
                continue
            val = row.iloc[c + 1]
            if val is None or (isinstance(val, float) and pd.isna(val)) \
                    or str(val).strip() == "":
                meta[label] = ""       # label present, value blank (distinct!)
            else:
                meta[label] = val.strip() if isinstance(val, str) else val
            anchor = r
            break
    return meta, anchor


# Enforces the metadata contract: the declared fields all become columns, the
# required ones must be present, "Report:" must equal the sheet name exactly,
# and any UNDECLARED label is a HIGH flag (that is how a typo gets caught
# instead of quietly creating a new column and drifting the schema).
def validate_metadata(meta, sheet_name, declared=DEFAULT_METADATA_FIELDS,
                      required=REQUIRED_METADATA_FIELDS):
    """Return (records, resolved) where `records` are report rows for any
    contract issue and `resolved` maps declared field -> value (pd.NA if absent).
    """
    recs = []
    by_norm = {_norm(k).rstrip(":").strip(): v for k, v in meta.items()}

    def rec(method, note, priority=REVIEW_HIGH, header=None, resolved=None):
        return {
            "source_header": header, "resolved_name": resolved,
            "method": method, "matched_against": None, "confidence": None,
            "review_priority": priority, "note": note,
        }

    resolved = {}
    for field in declared:
        key = _norm(field).rstrip(":").strip()
        val = by_norm.get(key, None)
        is_required = field in required

        if val is None:                      # label absent entirely
            resolved[field] = pd.NA
            if is_required:
                recs.append(rec(
                    "metadata_missing",
                    f"required metadata label '{field}:' not found above the "
                    f"pivot header", header=field, resolved=field,
                ))
            else:
                recs.append(rec(
                    "metadata_absent",
                    f"optional metadata '{field}:' not provided; column nulled",
                    priority=REVIEW_OK, header=field, resolved=field,
                ))
            continue

        if val == "":                        # label present, value blank
            resolved[field] = pd.NA
            recs.append(rec(
                "metadata_empty",
                f"metadata label '{field}:' found but the cell beside it is blank",
                priority=REVIEW_HIGH if is_required else REVIEW_OK,
                header=field, resolved=field,
            ))
            continue

        resolved[field] = val

    # Undeclared labels in the block -> HIGH. Catches typos and silent additions.
    declared_norm = {_norm(f).rstrip(":").strip() for f in declared}
    for key, val in by_norm.items():
        if key not in declared_norm:
            recs.append(rec(
                "metadata_unexpected",
                f"metadata label '{key}:' is not a declared field; declared are "
                f"{list(declared)}",
                header=key,
            ))

    # The one hard name constraint: Report: cell must equal the sheet name.
    report_name = resolved.get("Report")
    if isinstance(report_name, str) and report_name:
        if len(report_name) > MAX_SHEET_NAME_LEN:
            recs.append(rec(
                "report_name_too_long",
                f"report name is {len(report_name)} chars; max is "
                f"{MAX_SHEET_NAME_LEN} so the sheet name can mirror it exactly",
                header=report_name, resolved="Report",
            ))
        if sheet_name is not None and report_name != sheet_name:
            recs.append(rec(
                "report_name_mismatch",
                f"'{REPORT_LABEL}' cell is {report_name!r} but the sheet is "
                f"named {sheet_name!r}; they must match exactly",
                header=report_name, resolved="Report",
            ))
    return recs, resolved


# ============================================================================
# Header-name helpers
# ============================================================================

# Tidies a header into a plain, comparable form (lowercase, single spaces,
# optional "Sum of"-style prefix removed) so matching ignores cosmetic noise.
def _norm(s, strip_prefixes=()) -> str:
    """Normalize a header for comparison: lowercase, collapse whitespace,
    and optionally strip a leading aggregation prefix like 'Sum of '."""
    t = re.sub(r"\s+", " ", str(s).strip().lower())
    for p in strip_prefixes:
        if t.startswith(p + " "):
            t = t[len(p) + 1 :].strip()
            break
    return t


# Reads the alias dictionary you provide (in either direction) and turns it
# into fast lookup tables the matcher can use.
def _build_alias_index(aliases):
    """Accept aliases in either direction and return
    (canonical_by_norm, alias_pairs) where:
      - canonical_by_norm: {normalized_canonical: canonical_original_casing}
      - alias_pairs:       list of (normalized_alias, canonical) tuples

    Supported input shapes (auto-detected per entry):
      {canonical: [alias, alias, ...]}   # canonical -> acceptable names
      {canonical: alias}                 # single acceptable name / rename
      {alias: canonical}                 # "this header IS that canonical"
    A list/tuple/set value is always the acceptable-names form. A single-string
    value is read as {alias: canonical} (the intuitive "Row labels" -> "DEPT").
    """
    canonical_by_norm, alias_pairs = {}, []
    if not aliases:
        return canonical_by_norm, alias_pairs
    for key, val in aliases.items():
        if isinstance(val, (list, tuple, set)):
            canonical = key
            for a in val:
                alias_pairs.append((_norm(a), canonical))
        else:
            canonical = val
            alias_pairs.append((_norm(key), canonical))
        canonical_by_norm[_norm(canonical)] = canonical
    return canonical_by_norm, alias_pairs


# Removes an aggregation prefix ("Sum of ...") but keeps the original casing of
# what's left, so an unmatched value header still comes out clean (-> "FY26").
def _strip_prefix_keepcase(header, strip_prefixes) -> str:
    """Strip a leading aggregation prefix while preserving the remaining
    text's original casing (e.g. 'Sum of FY26' -> 'FY26')."""
    stripped = str(header).strip()
    low = stripped.lower()
    for p in strip_prefixes:
        if low.startswith(p + " "):
            return stripped[len(p) + 1 :].strip()
    return stripped


# Decides what ONE header should become: tries exact, then alias, then fuzzy,
# and records how sure it is (this is where a header earns an OK or HIGH flag).
def _resolve_one(header, canonical_by_norm, alias_pairs, fuzzy,
                 accept_threshold, review_threshold, strip_prefixes):
    """Resolve a single header, returning a report record (dict)."""
    hn = _norm(header, strip_prefixes)

    def rec(resolved, method, matched, conf, review, note=""):
        return {
            "source_header": header,
            "resolved_name": resolved,
            "method": method,
            "matched_against": matched,
            "confidence": conf,
            "review_priority": review,
            "note": note,
        }

    # 1: exact canonical
    if hn in canonical_by_norm:
        c = canonical_by_norm[hn]
        return rec(c, "canonical_exact", c, 1.0, REVIEW_OK)
    # 2: exact alias
    for n, c in alias_pairs:
        if n == hn:
            return rec(c, "alias_exact", n, 1.0, REVIEW_OK)
    # 3: fuzzy
    pool = [(n, canonical_by_norm[n]) for n in canonical_by_norm] + list(alias_pairs)
    if fuzzy and pool:
        best_n, best_c, best_r = None, None, 0.0
        for n, c in pool:
            r = SequenceMatcher(None, hn, n).ratio()
            if r > best_r:
                best_r, best_c, best_n = r, c, n
        if best_r >= accept_threshold:
            return rec(best_c, "fuzzy", best_n, round(best_r, 3), REVIEW_OK)
        if best_r >= review_threshold:
            return rec(
                best_c, "fuzzy_low_confidence", best_n, round(best_r, 3), REVIEW_HIGH,
                note="low-confidence fuzzy match (>= review_threshold); verify",
            )
        # 4: below review threshold -> do NOT rename; pass through. Only note
        #    the closest candidate when it's a genuine near-miss worth a look
        #    (avoids noisy notes on cleanly prefix-stripped value columns).
        cleaned = _strip_prefix_keepcase(header, strip_prefixes)
        note = ""
        if best_c is not None and best_r >= NEARMISS_NOTE_FLOOR:
            note = f"no confident match (closest '{best_c}' @ {round(best_r, 3)})"
        return rec(cleaned, "passthrough", None, None, REVIEW_OK, note=note)

    # No pool to match against -> passthrough (prefix-cleaned)
    return rec(_strip_prefix_keepcase(header, strip_prefixes),
               "passthrough", None, None, REVIEW_OK)


# Safety net: if two headers claim the same target name, lets the more-confident
# one win and sends the other back for review instead of silently merging them.
def _resolve_collisions(records):
    """When several headers resolve to the same name, keep the highest-
    confidence one and revert the others to a passthrough of their cleaned
    original header, flagging each reverted row HIGH for review. Mutates and
    returns the records list."""
    from collections import defaultdict

    by_name = defaultdict(list)
    for r in records:
        by_name[r["resolved_name"]].append(r)

    for name, group in by_name.items():
        if len(group) <= 1:
            continue
        # Winner = highest confidence (exact matches have confidence 1.0);
        # None confidence (passthrough) sorts lowest.
        winner = max(group, key=lambda r: (r["confidence"] or 0.0))
        for r in group:
            if r is winner:
                continue
            r["review_priority"] = REVIEW_HIGH
            r["note"] = (
                f"collision on '{name}' with "
                f"'{winner['source_header']}'; reverted to original name"
            ).strip()
            r["resolved_name"] = str(r["source_header"]).strip()
            r["method"] = "collision_reverted"
            r["matched_against"] = None
            r["confidence"] = None
    return records


# Runs the whole naming decision over a LIST of headers and returns one
# report row per header (the raw material for the kickback report).
def resolve_headers(headers, aliases=None, fuzzy=True,
                    accept_threshold=DEFAULT_ACCEPT_THRESHOLD,
                    review_threshold=DEFAULT_REVIEW_THRESHOLD,
                    strip_prefixes=DEFAULT_AGG_PREFIXES):
    """Resolve a list of raw headers to canonical names.

    Returns a list of report records (dicts), one per header, with keys:
      source_header, resolved_name, method, matched_against, confidence,
      review_priority, note.

    `method` is one of: canonical_exact, alias_exact, fuzzy,
    fuzzy_low_confidence, passthrough, collision_reverted.
    Rows needing attention have review_priority == "HIGH".
    """
    canonical_by_norm, alias_pairs = _build_alias_index(aliases)
    records = [
        _resolve_one(h, canonical_by_norm, alias_pairs, fuzzy,
                     accept_threshold, review_threshold, strip_prefixes)
        for h in headers
    ]
    return _resolve_collisions(records)


# Convenience shortcut: turn the report rows into a plain {old_name: new_name} map.
def mapping_from_records(records):
    """Convenience: {source_header: resolved_name} from report records."""
    return {r["source_header"]: r["resolved_name"] for r in records}


# ============================================================================
# Value healing — dimension MEMBERS (not headers)
# ============================================================================
#
# Two tiers, and the distinction is the whole point:
#
#   heal_universal  -- encoding noise. Cannot merge two members under ANY
#                      vocabulary, so it is always on and never asked about.
#   heal_licensed   -- conditionally safe. Only applied where the level's config
#                      licenses it AND the collision test proves the license does
#                      not merge two declared members.
#
# Anything that survives healing and still does not match a declared member is
# a DECISION, not a defect. It goes to a human. The parser never guesses a key.


# Encoding-level cleanup. Information-preserving: there is no vocabulary in
# which a non-breaking space or a zero-width joiner distinguishes two members.
def heal_universal(s) -> str:
    """Always-safe healing: unicode form, invisible characters, whitespace."""
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return s
    if not isinstance(s, str):
        # Numeric-coded members (cost centres, dept numbers) are legitimate.
        # Compare them as text so a declared "10" matches an int 10 -- but note
        # this never strips leading zeros, which ARE significant in coded dims.
        s = str(s)
    s = unicodedata.normalize("NFKC", s)   # NBSP -> space, full-width -> ASCII
    s = ZERO_WIDTH_RE.sub("", s)           # invisible chars: ZWSP/ZWNJ/ZWJ/BOM
    s = re.sub(r"\s+", " ", s).strip()     # collapse runs (incl. line breaks)
    return s


# Conditionally-safe healing. NOT applied unless the level's license allows it,
# and a license is only allowed once validate_license() proves it collides
# nothing in that level's declared vocabulary.
def heal_licensed(s, license=None) -> str:
    """Apply the healings this level has licensed (punctuation, case)."""
    if not isinstance(s, str):
        return s
    lic = {**DEFAULT_LICENSE, **(license or {})}
    if lic.get("punctuation"):
        s = PUNCT_RE.sub("", s)
    if lic.get("case"):
        s = s.casefold()
    return s


# The comparison key for a member: universal healing, then whatever this level
# licenses. Both the incoming value and the declared vocabulary go through this
# same function, so they are always compared on equal terms.
def member_key(s, license=None) -> str:
    """Normalized lookup key for a dimension member."""
    return heal_licensed(heal_universal(s), license)


# THE LICENSE CHECK. A license is a claim by the SME ("dashes do not matter for
# Service"). This proves the claim is self-consistent against their own declared
# vocabulary: if applying it would fold two declared members into one, the claim
# contradicts the dimension and we refuse to start. The machine does not decide
# what is ignorable -- it decides whether YOUR answer is coherent.
def validate_license(vocabulary, license) -> tuple:
    """Return (ok, collisions). A license is valid only if it never merges two
    declared members. Re-evaluated whenever the vocabulary grows, so the day a
    genuinely distinct 'F-1' is declared, punctuation-stripping stops being
    licensed for that level and the system knows without anyone remembering."""
    if not vocabulary:
        return True, {}
    seen, collisions = {}, {}
    for m in vocabulary:
        k = member_key(m, license)
        if k in seen and seen[k] != m:
            collisions.setdefault(k, {seen[k]}).add(m)
        seen[k] = m
    return (not collisions), {k: sorted(v) for k, v in collisions.items()}


# ============================================================================
# Configuration — the SME-owned declaration
# ============================================================================
#
# The spec is the artifact; this module is only the thing that obeys it. Loading
# validates the spec against ITSELF (every license vs. its own vocabulary) before
# a single file is opened, so a contradictory declaration fails at startup rather
# than halfway through a batch.


# Reads configuration.json and expands it into everything the parser needs:
# aliases, ordered levels, per-level vocabularies and normalization licenses.
def load_configuration(path=DEFAULT_CONFIG_PATH):
    """Load and validate the declaration. Raises on a self-contradictory spec.

    Returns a dict with:
      aliases           -> {canonical: [source names]}   (for resolve_headers)
      hierarchy_levels  -> [canonical, ...] in business order
      vocabularies      -> {canonical: set(members) | None}   None = open domain
      licenses          -> {canonical: {"punctuation": bool, "case": bool}}
      plus the metadata / matching / reconciliation blocks as given.
    """
    with open(path, "r", encoding="utf-8") as fh:
        cfg = json.load(fh)

    aliases, levels, vocabularies, licenses = {}, [], {}, {}
    modes, bootstrap, parents = {}, {}, {}
    prev_level = None
    for entry in cfg.get("hierarchy", []):
        canonical = entry["canonical"]
        levels.append(canonical)
        if entry.get("aliases"):
            aliases[canonical] = list(entry["aliases"])
        vocab = entry.get("vocabulary")
        vocabularies[canonical] = set(vocab) if vocab else None
        licenses[canonical] = {**DEFAULT_LICENSE, **(entry.get("normalization") or {})}
        bootstrap[canonical] = bool(entry.get("bootstrap", False))
        # The level directly above this one, used by registry mode to remember
        # which parent each member has been seen under.
        parents[canonical] = prev_level
        prev_level = canonical

        # GOVERNANCE MODE. Explicit, because "vocabulary is null" quietly meaning
        # "ungoverned" is exactly the kind of implicit rule that gets a level
        # left unchecked by accident.
        mode = entry.get("mode") or (MODE_CLOSED if vocab else MODE_OBSERVED)
        if mode not in VALID_MODES:
            raise ValueError(
                f"configuration error: level '{canonical}' has mode {mode!r}; "
                f"must be one of {VALID_MODES}."
            )
        if mode == MODE_CLOSED and not vocab:
            raise ValueError(
                f"configuration error: level '{canonical}' is mode 'closed' but "
                f"declares no vocabulary. A closed level with no members would "
                f"reject every file."
            )
        if mode != MODE_CLOSED and vocab:
            raise ValueError(
                f"configuration error: level '{canonical}' is mode {mode!r} but "
                f"also declares a vocabulary. Membership for this level comes "
                f"from the register, not the config -- remove one or the other."
            )
        if mode != MODE_REGISTRY and entry.get("bootstrap"):
            raise ValueError(
                f"configuration error: 'bootstrap' is only meaningful for mode "
                f"'registry'; level '{canonical}' is mode {mode!r}."
            )
        modes[canonical] = mode

        # Fail at startup, not mid-batch: prove the license does not merge two
        # declared members of THIS level. (For registry levels the same test is
        # re-run against the APPROVED REGISTER at parse time, so the day a
        # genuinely distinct 'F-1' is approved, punctuation-stripping stops being
        # licensed for that level and the system knows without anyone checking.)
        ok, collisions = validate_license(vocabularies[canonical], licenses[canonical])
        if not ok:
            raise ValueError(
                f"configuration error: normalization license for level "
                f"'{canonical}' would merge declared members {collisions}. "
                f"Either the members are the same thing (remove one) or the "
                f"license is wrong (tighten it)."
            )

    meta = cfg.get("metadata", {})
    match = cfg.get("matching", {})
    recon = cfg.get("reconciliation", {})
    pars = cfg.get("parsing", {})
    out = cfg.get("output", {})
    over = cfg.get("overrides", {})
    return {
        "version": cfg.get("version"),
        # ---- format semantics (US-locale strings; declared, not hardcoded) ----
        "strip_prefixes": tuple(pars.get("aggregation_prefixes",
                                         DEFAULT_AGG_PREFIXES)),
        "total_re": build_total_re(pars.get("total_labels")),
        "blank_literal": pars.get("blank_member_literal", BLANK_MEMBER_LITERAL),
        "metadata_scan_rows": pars.get("metadata_scan_rows", METADATA_SCAN_ROWS),
        "value_threshold": pars.get("value_column_threshold",
                                    DEFAULT_VALUE_THRESHOLD),
        "nearmiss_floor": pars.get("nearmiss_note_floor", NEARMISS_NOTE_FLOOR),
        # ---- output shape ----
        "var_name": out.get("var_name", "Period"),
        "value_name": out.get("value_name", "Value"),
        "sheet": out.get("sheet", 0),
        "allow_missing_levels": over.get("allow_missing_levels", {}),
        "aliases": aliases,
        "hierarchy_levels": levels,
        "vocabularies": vocabularies,
        "licenses": licenses,
        "modes": modes,
        "bootstrap": bootstrap,
        "parents": parents,
        "registry_path": cfg.get("registry", {}).get("path", DEFAULT_REGISTRY_PATH),
        "drift_tolerance": cfg.get("registry", {}).get(
            "cardinality_drift_tolerance", DEFAULT_DRIFT_TOLERANCE),
        "metadata_fields": tuple(meta.get("declared_fields", DEFAULT_METADATA_FIELDS)),
        "required_metadata": tuple(meta.get("required_fields", REQUIRED_METADATA_FIELDS)),
        "check_report_name": meta.get("check_report_name", True),
        "fuzzy": match.get("fuzzy", True),
        "accept_threshold": match.get("accept_threshold", DEFAULT_ACCEPT_THRESHOLD),
        "review_threshold": match.get("review_threshold", DEFAULT_REVIEW_THRESHOLD),
        "validate_totals": recon.get("validate_totals", True),
        "recon_tolerance": recon.get("tolerance", RECON_TOLERANCE),
    }


# Accepts the declaration however the caller has it: a path, an already-loaded
# dict, or nothing. Keeps every entry point config-aware without each one
# needing to know where the JSON lives.
def _expand_config(config):
    """Normalize `config` (path | dict | None) into a loaded config dict."""
    if config is None:
        return None
    if isinstance(config, dict):
        return config if "hierarchy_levels" in config else load_configuration(config)
    return load_configuration(config)


# ============================================================================
# Blank dimension cells — suppressed repeat vs. genuinely missing
# ============================================================================
#
# See the "Blank semantics" note in GLOBALS. The left-prefix property is what
# lets us decode a blank instead of guessing at it, with no declaration from the
# SME and no pivot-vs-flat sniffing: the blank tells you what it is by where it
# sits.


# True when a cell holds nothing at all (None / NaN / whitespace-only).
def _is_blank(v) -> bool:
    """Blank means absent: None, NaN, or a string with no visible content."""
    if v is None:
        return True
    if isinstance(v, float) and pd.isna(v):
        return True
    if isinstance(v, str) and not v.strip():
        return True
    return v is pd.NA


# Fills suppressed repeats and flags real holes. Level-aware on the outline
# column: a blank there is a repeat of the last label AT THE SAME INDENT LEVEL,
# never simply the cell above -- on an indented sheet the cell above is a CHILD,
# and a naive forward-fill would stamp "A subtype 2" onto A's siblings while the
# sums still tied perfectly. Silent-wrong, in exactly the way this module exists
# to prevent.
def resolve_blank_dimensions(body, dim_cols, levels=None):
    """Resolve blanks in the dimension columns. Returns (body, records).

    Left-contiguous blank prefix -> suppressed repeat -> filled (logged OK).
    Blank with a populated cell to its LEFT -> genuinely missing -> HIGH.
    """
    if not dim_cols:
        return body, []
    body = body.copy()
    n = len(body)
    levels = list(levels) if levels is not None else [0] * n
    recs, last_by_level, last_by_col = [], {}, {}
    filled_counts = {c: 0 for c in dim_cols}
    grid = {c: list(body[c]) for c in dim_cols}

    for i in range(n):
        row = [grid[c][i] for c in dim_cols]
        blanks = [j for j, v in enumerate(row) if _is_blank(v)]
        prefix = list(range(len(blanks)))

        if blanks and blanks == prefix:            # left prefix -> suppressed repeat
            for j in blanks:
                col = dim_cols[j]
                if j == 0:
                    fill = last_by_level.get(levels[i])   # same-level, not row above
                else:
                    fill = last_by_col.get(col)
                if fill is None:
                    recs.append(_member_rec(
                        DIM_MISSING_VALUE, col, None,
                        f"blank in '{col}' with no label above it to repeat",
                    ))
                else:
                    grid[col][i] = fill
                    filled_counts[col] += 1
        elif blanks:                               # not a prefix -> a real hole
            for j in blanks:
                col = dim_cols[j]
                left = dim_cols[j - 1] if j else None
                recs.append(_member_rec(
                    DIM_MISSING_VALUE, col, None,
                    f"blank in '{col}' but '{left}' is populated -- this is a "
                    f"missing value, not a suppressed repeat; not filled",
                ))

        # Remember the latest real label per column, and per indent level for the
        # outline column (the level bookkeeping is what makes nesting safe).
        for j, col in enumerate(dim_cols):
            v = grid[col][i]
            if not _is_blank(v):
                last_by_col[col] = v
                if j == 0:
                    last_by_level[levels[i]] = v

    for c in dim_cols:
        body[c] = grid[c]
        if filled_counts[c]:
            recs.append(_member_rec(
                LABEL_REPEAT_FILLED, c, None,
                f"{filled_counts[c]} suppressed repeat label(s) filled in '{c}' "
                f"from the last label at the same level",
                priority=REVIEW_OK,
            ))
    return body, recs


# ============================================================================
# The member registry (MODE_REGISTRY)
# ============================================================================
#
# Nobody hand-maintains a five-thousand-row dimension. What they maintain is a
# CONTROLLED PROCESS FOR ADDING ROWS TO ONE. That is the whole idea here:
#
#   configuration.json  = POLICY. Which levels exist, how each is governed.
#                         Small, stable, human-owned, reviewed with SMEs.
#   members.json        = THE REGISTER. Every member ever approved, per level.
#                         Large, churning, MACHINE-APPENDED and HUMAN-APPROVED.
#                         Never hand-edited.
#
# A new member is recorded as PENDING and blocks the load. Approval is BULK --
# "47 new Service codes this load: approve all / approve selected / reject" --
# so it is one decision, not forty-seven edits. Once approved, never asked again.
#
# Each entry also carries the parents it has been seen under. That is the seed of
# the only check we have that reaches into the misattribution blind spot: a
# member that has lived under DEPT B for six months and arrives today under DEPT
# A is not proof of an error, but it IS a signal, and the file alone can never
# provide one. History is the second source the spreadsheet cannot be.


# Reads (or creates) the register. Structure is deliberately boring so a human
# can read a diff of it in a PR.
def load_registry(path=DEFAULT_REGISTRY_PATH):
    """Load the member register. A missing file is an empty register, not an
    error -- the first run of a registry level has nothing to remember yet."""
    try:
        with open(path, "r", encoding="utf-8") as fh:
            reg = json.load(fh)
    except (FileNotFoundError, json.JSONDecodeError):
        reg = {}
    reg.setdefault("levels", {})
    return reg


# Writes the register back. Called explicitly by the caller, never as a hidden
# side effect of parsing -- a parse must not mutate governance state on its own.
def save_registry(registry, path=DEFAULT_REGISTRY_PATH):
    """Persist the member register."""
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(registry, fh, indent=2, sort_keys=True)
    return path


# The approved set for a level -- what MODE_REGISTRY checks against.
def approved_members(registry, level):
    """Members of `level` that a human has signed off on."""
    entries = (registry.get("levels", {}).get(level, {})).get("members", {})
    return {m for m, e in entries.items() if e.get("approved")}


# Everything seen but not yet signed off. This is the reviewer's worklist.
def pending_members(registry, level=None):
    """{level: [member, ...]} awaiting approval."""
    out = {}
    for lvl, blk in registry.get("levels", {}).items():
        if level and lvl != level:
            continue
        p = [m for m, e in blk.get("members", {}).items() if not e.get("approved")]
        if p:
            out[lvl] = sorted(p)
    return out


# BULK approval. The reason registry mode does not become a treadmill: the
# legitimate path through the gate has to be faster than the path around it, or
# somebody widens the vocabulary to a wildcard and you are back to six
# departments named "A".
def approve_members(registry, level, members=None, approver=None):
    """Approve pending members. `members=None` approves ALL pending for `level`."""
    blk = registry.setdefault("levels", {}).setdefault(level, {}).setdefault("members", {})
    targets = list(members) if members else [m for m, e in blk.items()
                                             if not e.get("approved")]
    stamp = pd.Timestamp.now("UTC").isoformat()
    for m in targets:
        e = blk.setdefault(m, {})
        e["approved"] = True
        e["approved_at"] = stamp
        e["approved_by"] = approver
    return sorted(targets)


# Records a sighting: registers a new member as pending (or auto-approves it at
# bootstrap) and remembers the parent it appeared under.
def _observe_member(registry, level, member, parent=None, bootstrap=False):
    """Record a sighting. Returns True if this member is NEW to the register."""
    blk = registry.setdefault("levels", {}).setdefault(level, {}).setdefault("members", {})
    is_new = member not in blk
    e = blk.setdefault(member, {})
    if is_new:
        e["first_seen"] = pd.Timestamp.now("UTC").isoformat()
        # BOOTSTRAP: on the first load of a granular level, EVERYTHING is new.
        # You cannot approve five thousand codes one at a time, and a reviewer
        # faced with five thousand HIGH rows disables the gate. So the first run
        # registers wholesale, loudly, and then bootstrap is turned off forever.
        e["approved"] = bool(bootstrap)
        if bootstrap:
            e["approved_at"] = e["first_seen"]
            e["approved_by"] = "bootstrap"
    if parent is not None and not _is_blank(parent):
        parents = e.setdefault("parents", {})
        parents[str(parent)] = parents.get(str(parent), 0) + 1
    return is_new


# ============================================================================
# Member-identity gate — THE KEY GATE
# ============================================================================
#
# The arithmetic gate asks "do the numbers tie?". The hierarchy gate asks "is the
# shape right?". Neither can see a row whose shape is perfect, whose numbers tie,
# and whose KEY is garbage -- "A " and "A" preserve every sum and are visually
# identical. This gate asks the remaining question: "is every key a member we
# declared?"
#
# Heal what is unambiguous. Fail on what is a decision. Nothing enters the
# warehouse as a key until it has been declared as one -- and that includes a
# genuinely new department, because new dimension membership is a business event
# that deserves a human's signature, not a cell's say-so.
#
# NOTE: fuzzy SUGGESTS here, it never APPLIES. A header is one decision per file
# and a wrong one is loud; a member is thousands of decisions per file and a
# wrong one is silent. Auto-accepting a fuzzy member match would manufacture the
# very orphan this gate exists to catch.


# Small helper so every record in this layer has the same shape as the rest of
# the kickback report.
def _member_rec(method, level, value, note, priority=REVIEW_HIGH,
                matched=None, confidence=None):
    """One report row for the healing / identity layer."""
    return {
        "source_header": value, "resolved_name": level, "method": method,
        "matched_against": matched, "confidence": confidence,
        "review_priority": priority, "note": note,
    }


# Canonicalizes every member against the declared vocabulary and reports anything
# it cannot resolve. Writes the CANONICAL value back into the frame -- the raw
# value survives in the report, so a heal is auditable but never invisible.
def validate_members(df, vocabularies=None, licenses=None,
                     blank_literal=BLANK_MEMBER_LITERAL, modes=None,
                     registry=None, bootstrap=None, parents=None,
                     drift_tolerance=DEFAULT_DRIFT_TOLERANCE):
    """Heal + identity-check the declared dimension columns.

    Dispatches on each level's GOVERNANCE MODE (see the MODE_* notes in GLOBALS):

      MODE_CLOSED    exact match against the hand-written vocabulary.
                     Undeclared -> HIGH.
      MODE_REGISTRY  exact match against the APPROVED register. A new member is
                     recorded as pending and -> HIGH, unless this level is
                     bootstrapping, in which case it is registered wholesale.
      MODE_OBSERVED  no membership check. Near-duplicate detection and
                     cardinality drift only.

    Returns (df, records). The CANONICAL value is written back to the frame; the
    raw value survives in the report, so a heal is auditable but never invisible.
    """
    vocabularies = vocabularies or {}
    licenses = licenses or {}
    modes = modes or {}
    bootstrap = bootstrap or {}
    parents = parents or {}
    df = df.copy()
    recs = []

    for level in (modes or vocabularies):
        if level not in df.columns:
            continue
        lic = licenses.get(level, DEFAULT_LICENSE)
        mode = modes.get(level) or (MODE_CLOSED if vocabularies.get(level)
                                    else MODE_OBSERVED)
        raw_values = [v for v in df[level].unique() if not _is_blank(v)]

        # ---- the approved set this level is checked against --------------
        if mode == MODE_CLOSED:
            vocab = vocabularies.get(level) or set()
        elif mode == MODE_REGISTRY:
            vocab = approved_members(registry or {}, level)
        else:
            vocab = None

        # ---- MODE_OBSERVED: no membership check, so look for the SHAPE of
        #      the error instead. At high cardinality this is the only control
        #      that can work -- nobody eyeballs five thousand project codes.
        if mode == MODE_OBSERVED:
            healed = {v: heal_universal(v) for v in raw_values}
            df[level] = df[level].map(lambda v: healed.get(v, v))
            fold = {"punctuation": True, "case": True}
            buckets = {}
            for v in set(healed.values()):
                buckets.setdefault(member_key(v, fold), set()).add(v)
            for group in buckets.values():
                if len(group) > 1:
                    recs.append(_member_rec(
                        MEMBER_NEAR_DUPLICATE, level, sorted(group)[0],
                        f"members {sorted(group)} in observed level '{level}' "
                        f"differ only by case/punctuation; they will split every "
                        f"rollup that touches them",
                    ))
            # Cardinality drift against the last load. Cheap, and it catches the
            # 3 -> 6 departments case without knowing what a department IS.
            recs += _check_cardinality_drift(
                registry, level, len(set(healed.values())), drift_tolerance)
            _remember_cardinality(registry, level, len(set(healed.values())))
            continue

        # ---- MODE_CLOSED / MODE_REGISTRY: membership is the identity check ---
        index = {member_key(m, lic): m for m in vocab}
        remap = {}
        for raw in raw_values:
            if heal_universal(raw).casefold() == str(blank_literal).casefold():
                remap[raw] = BLANK_MEMBER_SENTINEL
                recs.append(_member_rec(
                    MEMBER_BLANK_LITERAL, level, raw,
                    f"source rendered a null grouping key as "
                    f"'{blank_literal}'; mapped to {BLANK_MEMBER_SENTINEL!r}",
                    priority=REVIEW_OK,
                ))
                continue

            key = member_key(raw, lic)
            if key in index:
                canonical = index[key]
                remap[raw] = canonical
                if raw != canonical:
                    recs.append(_member_rec(
                        MEMBER_NORMALIZED, level, raw,
                        f"healed to declared member {canonical!r}",
                        priority=REVIEW_OK, matched=canonical, confidence=1.0,
                    ))
                continue

            # Not a known member. What happens next is the whole difference
            # between the two governed modes.
            healed_raw = heal_universal(raw)
            best, ratio = _nearest_member(healed_raw, vocab)
            hint = (f"; closest known member is {best!r} ({round(ratio, 3)})"
                    if best and ratio >= NEARMISS_NOTE_FLOOR else "")

            if mode == MODE_REGISTRY:
                boot = bool(bootstrap.get(level))
                _observe_member(registry, level, healed_raw, bootstrap=boot)
                remap[raw] = healed_raw
                if boot:
                    recs.append(_member_rec(
                        MEMBER_REGISTERED, level, raw,
                        f"{healed_raw!r} registered by BOOTSTRAP without review "
                        f"-- turn bootstrap off for '{level}' after this load",
                        priority=REVIEW_OK,
                    ))
                else:
                    recs.append(_member_rec(
                        MEMBER_PENDING, level, raw,
                        f"{healed_raw!r} is a NEW member of '{level}' and is "
                        f"pending approval{hint}; approve in bulk with "
                        f"approve_members()",
                        matched=best, confidence=round(ratio, 3) if best else None,
                    ))
            else:  # MODE_CLOSED -- a new member is a business DECISION
                recs.append(_member_rec(
                    MEMBER_UNDECLARED, level, raw,
                    f"{raw!r} is not a declared member of '{level}'{hint}"
                    f"{' -- approve or correct' if hint else ''}",
                    matched=best, confidence=round(ratio, 3) if best else None,
                ))

        if remap:
            df[level] = df[level].map(lambda v: remap.get(v, v))

        # ---- registry mode also REMEMBERS: sightings + the parent each member
        #      appeared under. That history is what later lets us notice a member
        #      that has quietly moved to a different parent -- the one signal we
        #      have that reaches into the misattribution blind spot.
        if mode == MODE_REGISTRY and registry is not None:
            parent_level = parents.get(level)
            recs += _track_parentage(df, registry, level, parent_level,
                                     bool(bootstrap.get(level)))

    return df, recs


# Nearest known member, folded fully so a case variant ('b' for 'B') still scores
# 1.0. The hint is for the REVIEWER -- it is never applied on their behalf.
def _nearest_member(value, vocab):
    """(closest_member, ratio) using a fully-folded comparison."""
    fold = {"punctuation": True, "case": True}
    k = member_key(value, fold)
    best, ratio = None, 0.0
    for m in vocab or ():
        r = SequenceMatcher(None, k, member_key(m, fold)).ratio()
        if r > ratio:
            best, ratio = m, r
    return best, ratio


# Records which parent each member appeared under, and flags a member that has
# moved. NOT proof of an error -- reorgs happen -- but the file alone can never
# tell you this, and history can.
def _track_parentage(df, registry, level, parent_level, bootstrap=False):
    recs = []
    if not parent_level or parent_level not in df.columns:
        return recs
    blk = registry.get("levels", {}).get(level, {}).get("members", {})
    pairs = df[[level, parent_level]].dropna().drop_duplicates()
    for member, parent in pairs.itertuples(index=False):
        known = set((blk.get(member, {}) or {}).get("parents", {}))
        if known and str(parent) not in known and not bootstrap:
            recs.append(_member_rec(
                MEMBER_PARENT_CHANGE, level, member,
                f"{member!r} has only ever been seen under {parent_level} "
                f"{sorted(known)} and now appears under {str(parent)!r}; confirm "
                f"this is a real move and not a misattribution",
            ))
        _observe_member(registry, level, member, parent=parent, bootstrap=True
                        if member in blk and blk[member].get("approved") else bootstrap)
    return recs


# Distinct-member count vs. the last load. Catches "3 departments became 6"
# without needing to know what a department is.
def _check_cardinality_drift(registry, level, n_now, tolerance):
    if not registry:
        return []
    prev = registry.get("levels", {}).get(level, {}).get("cardinality")
    if not prev or not n_now:
        return []
    change = abs(n_now - prev) / prev
    if change > tolerance:
        return [_member_rec(
            CARDINALITY_DRIFT, level, None,
            f"'{level}' had {prev} distinct members at the last load and has "
            f"{n_now} now ({change:.0%} change, tolerance {tolerance:.0%})",
        )]
    return []


def _remember_cardinality(registry, level, n):
    if registry is None:
        return
    registry.setdefault("levels", {}).setdefault(level, {})["cardinality"] = n


# ============================================================================
# Dimension-completeness gate
# ============================================================================


# Five lines that make the entire silent-null class impossible to pass. A row
# with a null key is worse than a wrong number: it drops out of every GROUP BY
# downstream and nobody ever sees it go.
def validate_dimension_completeness(df, hierarchy_levels):
    """Any null in a DECLARED hierarchy level is HIGH. Independent of how the
    null got there (suppressed repeat, bad alias, upstream hole)."""
    recs = []
    for level in hierarchy_levels or []:
        if level not in df.columns:
            continue
        n = int(df[level].isna().sum())
        if n:
            recs.append(_member_rec(
                DIMENSION_NULL, level, None,
                f"{n} row(s) have no value for declared level '{level}'; every "
                f"row must be addressable at every declared level",
            ))
    return recs


# ============================================================================
# Sheet-shape detection
# ============================================================================

# Answers "how numeric is this column?" as a 0..1 fraction — used to tell the
# data/value columns apart from the text label columns.
def _looks_numeric(series: pd.Series) -> float:
    """Fraction of non-null cells in `series` that are numeric."""
    non_null = series.dropna()
    if len(non_null) == 0:
        return 0.0
    def is_num(x):
        if isinstance(x, (int, float, np.integer, np.floating)) and not isinstance(x, bool):
            return True
        # numeric-looking strings e.g. "1,234"
        if isinstance(x, str):
            return bool(NUMERIC_RE.fullmatch(x))
        return False
    return non_null.map(is_num).mean()


# Figures out which row is the real header by finding where the numbers start
# and taking the label row just above them (skips junk rows at the very top).
def _find_header_row(raw: pd.DataFrame, min_value_cols: int = 1) -> int:
    """
    Find the index (0-based, within `raw`) of the header row.

    Heuristic: the header row is the label row directly above the first run of
    mostly-numeric rows (the pivot's value block), which lets us skip any
    metadata/junk rows sitting above the real table.
    """
    n_rows = raw.shape[0]

    def _is_numeric_cell(v) -> bool:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return False
        if isinstance(v, bool):
            return False
        if isinstance(v, (int, float, np.integer, np.floating)):
            return True
        if isinstance(v, str):
            return bool(NUMERIC_RE.fullmatch(v))
        return False

    def row_numeric_count(row) -> int:
        return sum(_is_numeric_cell(v) for v in row.values)

    # 1) Find the first row that starts a run of >=2 "mostly numeric" rows.
    #    That run is the pivot's data body; the header sits just above it.
    first_data = None
    for r in range(n_rows):
        if row_numeric_count(raw.iloc[r]) >= min_value_cols:
            nxt = raw.iloc[r + 1] if r + 1 < n_rows else None
            if nxt is None or row_numeric_count(nxt) >= min_value_cols:
                first_data = r
                break
    if first_data is None:
        raise ValueError("Could not locate a numeric data region.")

    # 2) Walk upward to the nearest non-empty, non-numeric row -> the header.
    for r in range(first_data - 1, -1, -1):
        row = raw.iloc[r]
        if row.notna().sum() >= 2 and _looks_numeric(row) < 0.5:
            return r
    raise ValueError("Could not locate a header row above the data region.")


# Resolves a raw header list to canonical column names once, so the two parsers
# and the reconciliation reader all name columns identically. Returns the pieces
# each caller needs: which source columns were labelled, their cleaned raw text,
# the per-column report records, and the resolved canonical names in order.
def _resolve_columns(headers, aliases, fuzzy, accept_threshold,
                     review_threshold, strip_prefixes):
    """Shared header-resolution boilerplate (was copy-pasted into extract_pivot,
    extract_hierarchical_pivot and read_total_row)."""
    valid_cols = [i for i, h in enumerate(headers) if pd.notna(h)]
    raw_names = [str(headers[i]).strip() for i in valid_cols]
    records = resolve_headers(
        raw_names, aliases=aliases, fuzzy=fuzzy,
        accept_threshold=accept_threshold, review_threshold=review_threshold,
        strip_prefixes=strip_prefixes,
    )
    col_names = [mapping_from_records(records)[n] for n in raw_names]
    return valid_cols, raw_names, records, col_names


# Splits resolved columns into value vs. dimension columns by numeric fraction,
# probing on the NON-total rows so a subtotal/total row can't skew the vote. This
# is the single definition the parsers and read_total_row now share -- previously
# three hand-kept-in-sync copies, and the reconciliation baseline depends on it
# classifying columns exactly as the parser did.
def _split_value_dims(body, col_names, total_re, value_threshold):
    """Return (value_cols, dim_cols) for a body whose columns are `col_names`."""
    probe = body[~body.apply(
        lambda r: any(isinstance(v, str) and total_re.search(v) for v in r.values),
        axis=1)]
    value_cols, dim_cols = [], []
    for c in col_names:
        (value_cols if _looks_numeric(probe[c]) >= value_threshold
         else dim_cols).append(c)
    return value_cols, dim_cols


# ============================================================================
# Main extraction + batch ingest
# ============================================================================

# The main event: reads ONE pivot sheet, cleans it up, renames the columns via
# the alias/fuzzy logic, and hands back tidy long-format rows (+ optional report).
def extract_pivot(
    path: str,
    sheet=0,
    value_name: str = "Value",
    var_name: str = "Period",
    aliases=None,
    fuzzy: bool = True,
    accept_threshold: float = DEFAULT_ACCEPT_THRESHOLD,
    review_threshold: float = DEFAULT_REVIEW_THRESHOLD,
    strip_prefixes=DEFAULT_AGG_PREFIXES,
    return_report: bool = False,
    vocabularies=None,
    licenses=None,
    total_re=None,
    value_threshold: float = DEFAULT_VALUE_THRESHOLD,
    blank_literal=BLANK_MEMBER_LITERAL,
    modes=None,
    registry=None,
    bootstrap=None,
    parents=None,
    drift_tolerance: float = DEFAULT_DRIFT_TOLERANCE,
):
    """
    Read a pivot-style sheet and return a tidy long-format DataFrame.

    Parameters
    ----------
    path : str
        Path to the .xlsx / .xls file.
    sheet : int | str
        Sheet index or name (e.g. "Sheet3").
    value_name, var_name : str
        Names for the value column and the un-pivoted header column.
    aliases : dict | None
        Header mapping, either direction per entry:
          {"DEPT": ["Row labels", "Department"]}  # canonical -> acceptable names
          {"Row labels": "DEPT"}                   # this header IS that canonical
    fuzzy : bool
        Enable fuzzy fallback matching.
    accept_threshold : float
        Fuzzy similarity at/above which a match is applied silently (OK).
    review_threshold : float
        Fuzzy similarity at/above which a match is still applied but flagged
        HIGH for review. Below this, the header is NOT renamed (passthrough).
    strip_prefixes : tuple[str]
        Aggregation prefixes stripped before matching and from unmatched
        headers ("Sum of FY26" -> "FY26"). Pass () to disable.
    return_report : bool
        If True, return (tidy_df, report_df). The report has one row per source
        column; rows with review_priority == "HIGH" need a human look.

    Returns
    -------
    pd.DataFrame, or (df, report_df) when return_report=True.
    """
    # Read the whole sheet with no header so we can find it ourselves.
    raw = pd.read_excel(path, sheet_name=sheet, header=None, dtype=object)
    # Drop entirely-empty trailing columns/rows that pandas sometimes adds.
    raw = raw.dropna(axis=1, how="all").dropna(axis=0, how="all").reset_index(drop=True)

    hdr = _find_header_row(raw)
    headers = raw.iloc[hdr].tolist()
    body = raw.iloc[hdr + 1 :].reset_index(drop=True)
    body.columns = range(body.shape[1])

    # Resolve headers -> canonical names (aliases + fuzzy + prefix cleanup),
    # collecting a per-column report. Collisions are logged & auto-reverted
    # rather than raised, so mass ingestion never stops on one bad file.
    valid_cols, raw_names, records, col_names = _resolve_columns(
        headers, aliases, fuzzy, accept_threshold, review_threshold,
        strip_prefixes)
    body = body[valid_cols]
    body.columns = col_names

    report = pd.DataFrame.from_records(records, columns=REPORT_COLUMNS)

    # Classify columns first, so the total test can be anchored to the outline
    # (first dimension) column rather than scanned across the whole row.
    total_re = total_re or TOTAL_RE
    value_cols, dim_cols = _split_value_dims(body, col_names, total_re,
                                             value_threshold)

    if not value_cols:
        raise ValueError("No numeric value columns detected to unpivot.")
    outline = dim_cols[0] if dim_cols else None

    # Drop total/subtotal rows (anchored) and blank rows.
    body = body[~body.apply(_is_total_row, axis=1, dim_cols=dim_cols,
                            outline=outline, total_re=total_re)]
    body = body.dropna(how="all").reset_index(drop=True)

    # Coerce value columns to numbers ("1,234", "(948)" -> -948, "$1,234").
    for c in value_cols:
        body[c] = to_numeric_column(body[c])

    # Drop rows where every dimension cell is blank (stray totals w/o label).
    if dim_cols:
        body = body[body[dim_cols].notna().any(axis=1)].reset_index(drop=True)

    # THE BODY ENDS WHERE THE NUMBERS END. Everything below the header was read,
    # which sweeps up the trailing notes/helper blocks these templates carry. A
    # row with no value in ANY period is not a data row -- cut it here, before
    # the member gate sees "*NOTES ..." and reports it as an undeclared member.
    body = body.dropna(subset=value_cols, how="all").reset_index(drop=True)

    # ORDER MATTERS. Heal + canonicalize members FIRST, so that the repeat-fill
    # below copies CANONICAL labels rather than raw ones; then resolve blanks.
    body, member_recs = validate_members(
        body, vocabularies, licenses, blank_literal, modes=modes,
        registry=registry, bootstrap=bootstrap, parents=parents,
        drift_tolerance=drift_tolerance,
    )
    body, blank_recs = resolve_blank_dimensions(body, dim_cols)
    if member_recs or blank_recs:
        report = pd.concat(
            [report, pd.DataFrame(member_recs + blank_recs, columns=REPORT_COLUMNS)],
            ignore_index=True,
        )

    tidy = body.melt(
        id_vars=dim_cols,
        value_vars=value_cols,
        var_name=var_name,
        value_name=value_name,
    )
    tidy = tidy.dropna(subset=[value_name]).reset_index(drop=True)
    if return_report:
        return tidy, report
    return tidy


# Runs extract_pivot over MANY files at once, stacking the results and reports;
# a file that blows up is logged as HIGH and skipped so the batch never stops.
def ingest(
    paths,
    sheet=None,
    value_name: str = None,
    var_name: str = None,
    aliases=None,
    fuzzy: bool = True,
    accept_threshold: float = DEFAULT_ACCEPT_THRESHOLD,
    review_threshold: float = DEFAULT_REVIEW_THRESHOLD,
    strip_prefixes=None,
    parser=None,
    hierarchy_levels=None,
    hierarchy=None,
    allow_missing_levels=None,
    check_report_name: bool = True,
    metadata_fields=DEFAULT_METADATA_FIELDS,
    required_metadata=REQUIRED_METADATA_FIELDS,
    validate_totals: bool = True,
    recon_tolerance: float = RECON_TOLERANCE,
    config=None,
    vocabularies=None,
    licenses=None,
    gate: bool = True,
    registry=None,
    registry_path=None,
    save_registry_on_exit: bool = True,
):
    """Mass-ingest many pivot sheets into one tidy DataFrame plus one combined
    kickback report.

    config : path to configuration.json (or a loaded dict). The one artifact the
        SMEs maintain: aliases, ordered levels, member vocabularies and
        normalization licenses. Validated against itself at load, so a
        self-contradictory declaration fails before a file is opened.

    gate : run load_gate() before returning (DEFAULT TRUE). A safety property you
        have to remember to invoke is not a safety property -- ANY HIGH row raises
        here. Pass gate=False only to inspect a failing batch's report.

    allow_missing_levels : per-report override for legitimately absent hierarchy
        levels. Scoped deliberately -- a blanket switch would also forgive real
        mapping mistakes. Two accepted shapes:
            {"F-DEPT-VAR-TEAM-SPRL": ["Service"]}   # keyed by report/sheet name
            {"Sample_5.xlsx": ["Service"]}          # keyed by file basename
        A missing level NOT listed for that report stays a HIGH flag.

    parser : which parse class to use per file. None -> parse (auto-routes flat
    vs hierarchical, enforces schema + name contract).

    Never raises on a single bad file: a file that fails to parse is recorded
    as a HIGH-priority row in the report (method='file_error') and skipped, so
    the batch always completes.

    Returns
    -------
    (data, report) : (pd.DataFrame, pd.DataFrame)
        data   -> concatenated tidy rows, with a leading 'source_file' column.
        report -> concatenated per-column report, leading 'source_file' column.
                  Filter review_priority == 'HIGH' for the items to eyeball.
    """
    import os

    # Expand the declaration once for the whole batch, not once per file.
    cfg = _expand_config(config)
    if cfg:
        aliases = {**cfg["aliases"], **(aliases or {})}
        hierarchy_levels = hierarchy_levels or cfg["hierarchy_levels"]
        vocabularies = vocabularies or cfg["vocabularies"]
        licenses = licenses or cfg["licenses"]
        metadata_fields = cfg["metadata_fields"]
        required_metadata = cfg["required_metadata"]
        check_report_name = cfg["check_report_name"]
        validate_totals = cfg["validate_totals"]
        recon_tolerance = cfg["recon_tolerance"]
        fuzzy = cfg["fuzzy"]
        accept_threshold = cfg["accept_threshold"]
        review_threshold = cfg["review_threshold"]
        strip_prefixes = strip_prefixes or cfg["strip_prefixes"]
        var_name = var_name or cfg["var_name"]
        value_name = value_name or cfg["value_name"]
        sheet = cfg["sheet"] if sheet is None else sheet
        # Per-report permitted-missing levels are part of the DECLARATION too:
        # an exception someone approved, recorded where everything else is.
        allow_missing_levels = allow_missing_levels or cfg["allow_missing_levels"]
        # Load the register ONCE for the batch, so members first seen in file 3
        # are already known by file 20 and are reported once, not seventeen times.
        registry_path = registry_path or cfg["registry_path"]
        if registry is None and MODE_REGISTRY in set(cfg["modes"].values()):
            registry = load_registry(registry_path)

    sheet = 0 if sheet is None else sheet
    var_name = var_name or "Period"
    value_name = value_name or "Value"
    strip_prefixes = strip_prefixes or DEFAULT_AGG_PREFIXES

    parse_fn = parser or parse
    overrides = allow_missing_levels or {}
    frames, reports = [], []
    stamp = pd.Timestamp.now("UTC").isoformat()
    for path in paths:
        name = os.path.basename(path)
        try:
            digest = file_digest(path)
        except Exception:  # noqa: BLE001 - a missing file is reported below
            digest = None
        # Resolve this file's permitted-missing levels: look up by sheet/report
        # name first, then by file basename. Default () == nothing permitted.
        try:
            key = _sheet_name_of(path, sheet)
        except Exception:  # noqa: BLE001
            key = None
        allowed = overrides.get(key, overrides.get(name, ()))

        # Only the default parse router understands the extra contract args.
        extra = ({"hierarchy_levels": hierarchy_levels, "hierarchy": hierarchy,
                  "allow_missing_levels": allowed,
                  "check_report_name": check_report_name,
                  "metadata_fields": metadata_fields,
                  "required_metadata": required_metadata,
                  "validate_totals": validate_totals,
                  "recon_tolerance": recon_tolerance,
                  "vocabularies": vocabularies,
                  "licenses": licenses,
                  "config": cfg,
                  "registry": registry}
                 if parser is None else {})
        try:
            df, rep = parse_fn(
                path, sheet=sheet, value_name=value_name, var_name=var_name,
                aliases=aliases, fuzzy=fuzzy, accept_threshold=accept_threshold,
                review_threshold=review_threshold, strip_prefixes=strip_prefixes,
                return_report=True, **extra,
            )
            # PROVENANCE_COLUMNS is the single declaration of these names/order.
            for i, (col, val) in enumerate(
                    zip(PROVENANCE_COLUMNS, (name, digest, stamp))):
                df.insert(i, col, val)
            df = add_row_keys(df, hierarchy_levels, var_name, value_name)
            rep.insert(0, "source_file", name)
            frames.append(df)
            reports.append(rep)
        except Exception as e:  # noqa: BLE001 - we want to log, not crash
            reports.append(pd.DataFrame([{
                "source_file": name,
                "source_sha256": digest,
                "source_header": None,
                "resolved_name": None,
                "method": "file_error",
                "matched_against": None,
                "confidence": None,
                "review_priority": REVIEW_HIGH,
                "note": f"{type(e).__name__}: {e}",
            }]))

    data = (pd.concat(frames, ignore_index=True)
            if frames else pd.DataFrame())
    report = (pd.concat(reports, ignore_index=True)
              if reports else pd.DataFrame())

    # Persist what the register LEARNED -- new members (pending or bootstrapped)
    # and the parents each was seen under. Written here, once, at batch level:
    # a parse must never mutate governance state as a hidden side effect.
    #
    # NOTE the deliberate ordering: we save BEFORE the gate raises. A blocked
    # batch has still taught us what the new members ARE, and that list is
    # exactly what the reviewer needs in order to unblock it. Losing it on the
    # exception would mean re-running the batch just to rediscover the worklist.
    if registry is not None and save_registry_on_exit and registry_path:
        save_registry(registry, registry_path)

    # FAIL-CLOSED BY DEFAULT. The gate used to be a function the caller had to
    # remember to call, which meant one forgotten line put un-reviewed data into
    # the warehouse. It is now a stage of the pipeline.
    if gate and len(report):
        load_gate(report, raise_on_fail=True)
    return data, report


# Filters a report down to just the rows a human should double-check.
def review_items(report):
    """Return just the HIGH-priority rows from a report."""
    return report[report["review_priority"] == REVIEW_HIGH].reset_index(drop=True)


# Shows the levels that were nulled by an explicit override -- visible, not
# silenced. These are accepted exceptions, NOT failures.
def override_items(report):
    """Return the OVERRIDDEN rows from a report (permitted-missing levels)."""
    return report[
        report["review_priority"] == REVIEW_OVERRIDDEN
    ].reset_index(drop=True)


# THE GATE: fail-closed. Any HIGH row stops the load. Overrides do not.
# Call this before writing anything downstream.
def load_gate(report, raise_on_fail: bool = True):
    """Decide whether a batch may load. Returns (ok, high_rows).

    Fail-closed: ANY HIGH row (missing level, name mismatch, collision,
    file error, low-confidence fuzzy match) blocks the load. Rows that were
    explicitly permitted appear as OVERRIDDEN and do NOT block.
    """
    high = review_items(report)
    ok = len(high) == 0
    if not ok and raise_on_fail:
        summary = high.groupby("method").size().to_dict()
        raise ValueError(
            f"Load blocked: {len(high)} HIGH-priority issue(s) {summary}. "
            "Review the report, fix the mapping, or explicitly permit the "
            "level via allow_missing_levels."
        )
    return ok, high


# ============================================================================
# Hierarchical (indented sub-level) pivot parsing  ---  second parse "class"
# ============================================================================
#
# Some pivots show an outline: a parent row (a subtotal) with its children
# indented beneath it in the SAME column. Real Excel PivotTables encode that
# indentation as the cell's alignment `indent` property (an integer outline
# level), NOT as characters in the text -- so the value is just "A subtype 1"
# with indent=1. Copy/pasted or hand-built sheets sometimes fake it with
# leading spaces instead. The reader below captures BOTH signals so either
# works, preferring the real alignment property when present.


# The actual openpyxl read, memoized on file IDENTITY (path + mtime + size), so
# the ~4 places that need the sheet within a single parse()/reconcile pass read
# and materialize it ONCE instead of hitting disk each time. The mtime+size in
# the key mean a re-saved file is a cache miss and gets re-read -- the cache can
# never serve stale bytes. Small maxsize: within a batch only the current file's
# reads need to stay warm; older files evict.
@functools.lru_cache(maxsize=8)
def _read_sheet_cached(abspath, sheet, _mtime_ns, _size):
    from openpyxl import load_workbook

    wb = load_workbook(abspath, data_only=True)
    try:
        ws = wb.worksheets[sheet] if isinstance(sheet, int) else wb[sheet]
        values, indents = [], []
        for row in ws.iter_rows():
            values.append([c.value for c in row])
            indents.append([int(c.alignment.indent or 0) for c in row])
    finally:
        wb.close()

    vdf = pd.DataFrame(values)
    idf = pd.DataFrame(indents).reindex(columns=vdf.columns, fill_value=0)
    # Drop columns that are entirely empty (keeps column positions aligned).
    keep = [c for c in vdf.columns if vdf[c].notna().any()]
    return vdf[keep].reset_index(drop=True), idf[keep].reset_index(drop=True)


# Reads a sheet with openpyxl so we can see cell indentation (which pandas
# throws away). Returns the values as a DataFrame plus a matching grid of
# indent levels, one per cell. Backed by the identity-keyed cache above; callers
# get fresh COPIES so nobody can mutate another caller's view.
def _read_sheet_with_indent(path, sheet=0):
    """Return (values_df, indent_df) where indent_df holds each cell's Excel
    alignment indent level (int). Column positions line up between the two."""
    st = os.stat(path)   # raises FileNotFoundError for a missing file, as before
    vdf, idf = _read_sheet_cached(os.path.abspath(path), sheet,
                                  st.st_mtime_ns, st.st_size)
    return vdf.copy(), idf.copy()


# Works out how deeply each row is indented in the outline column, using the
# real Excel indent if the sheet has any, otherwise counting leading spaces.
# Distinct indent widths are ranked into clean levels 0, 1, 2, ...
def _indent_levels(label_values, label_indents):
    """Given the outline column's values and raw indent attrs, return a list of
    normalized integer levels (0 = top) and the cleaned (un-indented) labels."""
    raw = []
    use_attr = any(int(x or 0) > 0 for x in label_indents)
    for val, attr in zip(label_values, label_indents):
        if use_attr:
            raw.append(int(attr or 0))
        else:
            s = val if isinstance(val, str) else ""
            raw.append(len(s) - len(s.lstrip(" ")))  # leading-space count

    # Rank the distinct indent widths so any unit (7 spaces, 1 attr, ...) -> 0,1,2
    distinct = sorted({r for r in raw})
    rank = {width: i for i, width in enumerate(distinct)}
    levels = [rank[r] for r in raw]

    cleaned = [v.strip() if isinstance(v, str) else v for v in label_values]
    return levels, cleaned


# Walks the outline top-to-bottom keeping a stack of ancestors, so each row
# learns its immediate parent (the join key) and its full path from the root.
def _hierarchy_lineage(labels, levels):
    """Given per-row outline labels and their integer levels (already computed),
    return (parent_labels, paths):
      parent_labels[i] -> the label of row i's immediate parent, or None at top.
      paths[i]         -> 'Root > ... > Self' ancestry string for row i.
    Works for arbitrary depth via an ancestor stack."""
    parent_labels, paths, stack = [], [], []  # stack of (level, label)
    for lbl, lv in zip(labels, levels):
        while stack and stack[-1][0] >= lv:
            stack.pop()
        parent_labels.append(stack[-1][1] if stack else None)
        paths.append(" > ".join([a[1] for a in stack] + [str(lbl)]))
        stack.append((lv, lbl))
    return parent_labels, paths


# The hierarchical counterpart to extract_pivot: same tidy long output, but it
# keeps parent (subtotal) rows AND their indented children, tagging each row
# with level / is_parent / is_leaf so ETL can filter or roll up as needed.
def extract_hierarchical_pivot(
    path: str,
    sheet=0,
    value_name: str = "Value",
    var_name: str = "Period",
    hierarchy_col=None,
    aliases=None,
    fuzzy: bool = True,
    accept_threshold: float = DEFAULT_ACCEPT_THRESHOLD,
    review_threshold: float = DEFAULT_REVIEW_THRESHOLD,
    strip_prefixes=DEFAULT_AGG_PREFIXES,
    return_report: bool = False,
    vocabularies=None,
    licenses=None,
    total_re=None,
    value_threshold: float = DEFAULT_VALUE_THRESHOLD,
    blank_literal=BLANK_MEMBER_LITERAL,
    modes=None,
    registry=None,
    bootstrap=None,
    parents=None,
    drift_tolerance: float = DEFAULT_DRIFT_TOLERANCE,
):
    """
    Parse an *indented* pivot (parent subtotal rows with children nested under
    them in one column) into tidy long form, preserving the hierarchy.

    Adds these columns to the usual output:
      sub_level : int, 0 = top-level, 1 = first indented sub-level, ...
                  (this is WITHIN-COLUMN indent depth, NOT the business
                  hierarchy level -- the business levels are the ordered
                  dimension columns themselves, e.g. DEPT > Service > Variant.)
      is_parent : True if the row is a subtotal that has children beneath it.
      is_leaf   : True if the row has no children (the additive detail rows).

    NOTE ON SUMMING: parents are subtotals of their children, so both live in the
    output. Sum over is_leaf == True for an additive rollup; parents are there
    for reconciliation, not for adding on top of their children.

    hierarchy_col : the outline column. None = auto (first dimension column).
    Other parameters behave exactly as in extract_pivot.
    """
    raw, ind = _read_sheet_with_indent(path, sheet)

    # Locate header row (reuses the flat detector) and slice the body + indents.
    hdr = _find_header_row(raw)
    headers = raw.iloc[hdr].tolist()
    body = raw.iloc[hdr + 1 :].reset_index(drop=True)
    body_ind = ind.iloc[hdr + 1 :].reset_index(drop=True)
    body.columns = range(body.shape[1])
    body_ind.columns = range(body_ind.shape[1])

    # Keep only labeled columns; resolve header names (aliases + fuzzy).
    valid_cols, raw_names, records, col_names = _resolve_columns(
        headers, aliases, fuzzy, accept_threshold, review_threshold,
        strip_prefixes)
    body = body[valid_cols]
    body_ind = body_ind[valid_cols]
    body.columns = col_names
    body_ind.columns = col_names
    report = pd.DataFrame.from_records(records, columns=REPORT_COLUMNS)

    # Classify value vs dimension columns (value cols are mostly numeric). We
    # look at rows that aren't the total row so subtotals don't skew it.
    total_re = total_re or TOTAL_RE
    value_cols, dim_cols = _split_value_dims(body, col_names, total_re,
                                             value_threshold)
    if not value_cols:
        raise ValueError("No numeric value columns detected to unpivot.")
    if not dim_cols:
        raise ValueError("No dimension columns found for the hierarchy.")

    # Pick the outline column: explicit arg, else the first dimension column.
    if hierarchy_col is None:
        hier = dim_cols[0]
    else:
        hier = mapping_from_records(records).get(hierarchy_col, hierarchy_col)
        if hier not in dim_cols:
            raise ValueError(f"hierarchy_col '{hier}' is not a dimension column.")

    # Compute indent levels + cleaned labels for the outline column, then write
    # the cleaned labels back so the value carries no leading whitespace.
    levels, cleaned = _indent_levels(list(body[hier]), list(body_ind[hier]))
    body[hier] = cleaned
    body["sub_level"] = levels

    # Drop the grand-total row (anchored to the outline column) and blank rows.
    mask_total = body.apply(_is_total_row, axis=1, dim_cols=dim_cols,
                            outline=hier, total_re=total_re)
    mask_keep = ~mask_total & body[dim_cols].notna().any(axis=1)
    body = body[mask_keep].reset_index(drop=True)

    # Coerce values first so the body can be cut at the last row that actually
    # carries a number (see the note in extract_pivot).
    for c in value_cols:
        body[c] = to_numeric_column(body[c])
    body = body.dropna(subset=value_cols, how="all").reset_index(drop=True)

    # ORDER MATTERS, AND THIS IS THE ORDER.
    #   1. indent already extracted above -- leading spaces ARE the hierarchy, so
    #      they must be consumed BEFORE any strip()/heal touches the labels. Do
    #      not "tidy" the healing earlier in this function: it flattens sub_level
    #      to 0 for every row and silently double-counts parents with children.
    #   2. heal + canonicalize members, so the repeat-fill copies CANONICAL labels
    #      and the lineage below is built from canonical ones.
    #   3. resolve blanks (suppressed repeats vs. real holes).
    #   4. only then compute parentage and paths.
    body, member_recs = validate_members(
        body, vocabularies, licenses, blank_literal, modes=modes,
        registry=registry, bootstrap=bootstrap, parents=parents,
        drift_tolerance=drift_tolerance,
    )
    body, blank_recs = resolve_blank_dimensions(
        body, dim_cols, levels=body["sub_level"].tolist()
    )
    if member_recs or blank_recs:
        report = pd.concat(
            [report, pd.DataFrame(member_recs + blank_recs, columns=REPORT_COLUMNS)],
            ignore_index=True,
        )

    # A row is a parent (subtotal) if the row immediately below it sits deeper
    # in the outline. Everything else is a leaf.
    lv = body["sub_level"].tolist()
    is_parent = [
        (i + 1 < len(lv) and lv[i + 1] > lv[i]) for i in range(len(lv))
    ]
    body["is_parent"] = is_parent
    body["is_leaf"] = [not p for p in is_parent]

    # Immediate parent (join key for rolling children back up) + full path.
    parent_labels, paths = _hierarchy_lineage(list(body[hier]), lv)
    body["parent_label"] = parent_labels
    body["path"] = paths

    id_cols = dim_cols + ["sub_level", "is_parent", "is_leaf", "parent_label", "path"]
    tidy = body.melt(
        id_vars=id_cols, value_vars=value_cols,
        var_name=var_name, value_name=value_name,
    )
    tidy = tidy.dropna(subset=[value_name]).reset_index(drop=True)

    if return_report:
        return tidy, report
    return tidy


# ============================================================================
# Auto-dispatch  ---  one entry point that sniffs the sheet and routes
# ============================================================================
#
# Design note: flat and hierarchical parsing share the same core (header
# detection, alias/fuzzy resolution, total-row handling, melting, reporting).
# Rather than maintain two scripts, we keep ONE module with two entry points and
# this thin router. It peeks at the outline column once; if any row is indented
# (real Excel indent property OR leading spaces) it uses the hierarchical
# parser, otherwise the flat one -- so callers can just say parse(path).


# Cheap check: does the outline (first dimension) column contain any indentation?
# Decides which parser the sheet needs.
def _sheet_is_hierarchical(path, sheet=0):
    """True if the outline column shows any indent signal (native indent attr
    or leading spaces), meaning the sheet has nested sub-levels."""
    try:
        raw, ind = _read_sheet_with_indent(path, sheet)
        hdr = _find_header_row(raw)
        body = raw.iloc[hdr + 1 :].reset_index(drop=True)
        body_ind = ind.iloc[hdr + 1 :].reset_index(drop=True)
        first_col = body.columns[0]
        # native Excel indent anywhere?
        if (body_ind[first_col].fillna(0).astype(int) > 0).any():
            return True
        # leading-space indent anywhere?
        lead = body[first_col].map(
            lambda v: (len(v) - len(v.lstrip(" "))) if isinstance(v, str) else 0
        )
        return bool((lead > 0).any())
    except Exception:
        return False  # if in doubt, let the flat parser try


# Guarantees every output has the same columns in the same order: flat results
# get the hierarchy columns filled in as level-0 leaves so flat and hierarchical
# frames stack cleanly in one batch.
def _harmonize(df, var_name, value_name):
    """Ensure df carries all HIERARCHY_COLUMNS (adding sane defaults for flat
    data) and put columns in a stable order: dimensions, hierarchy metadata,
    then the period + value columns last."""
    df = df.copy()
    lead = [c for c in df.columns
            if c not in HIERARCHY_COLUMNS + [var_name, value_name]]
    defaults = {
        "sub_level": 0, "is_parent": False, "is_leaf": True,
        "parent_label": pd.NA, "path": pd.NA,
    }
    for col, default in defaults.items():
        if col not in df.columns:
            # For flat data, seed 'path' with the row's top-level dimension so
            # to_level_columns() still yields a Level_0 (uniform with nested data).
            if col == "path" and lead:
                # astype("string") turns a genuine NaN into the literal text
                # "nan", which then lands in the warehouse as three characters
                # that pass every IS NOT NULL check. Keep nulls as nulls.
                df[col] = df[lead[0]].where(df[lead[0]].notna(), pd.NA).astype("string")
            else:
                df[col] = default
    return df[lead + HIERARCHY_COLUMNS + [var_name, value_name]]


# Serving-layer convenience: explodes the 'path' string into Level_0, Level_1,
# ... columns for tools (e.g. Power BI matrix) that drill down on columns rather
# than a parent-child reference. A derived view, not the canonical shape.
def to_level_columns(df, prefix="Level_", path_col="path"):
    """Return a copy of df with the ancestry in `path_col` split into
    prefix+0, prefix+1, ... columns (forward-filled down the path). Rows with no
    path (flat/level-0) keep their own label in Level_0 where available."""
    out = df.copy()
    if path_col not in out.columns:
        return out
    parts = out[path_col].map(
        lambda p: p.split(" > ") if isinstance(p, str) else []
    )
    depth = int(parts.map(len).max() or 0)
    for i in range(depth):
        out[f"{prefix}{i}"] = parts.map(lambda lst: lst[i] if i < len(lst) else pd.NA)
    return out


# Turns ONE upfront business-hierarchy spec (the artifact you build with SMEs)
# into the two things the parser needs: the ordered canonical level names and
# the alias map. Define the hierarchy once; pass it in; done.
def build_hierarchy(spec):
    """Expand a business-hierarchy spec into (aliases, hierarchy_levels).

    `spec` defines the fixed business hierarchy in order (level 1, 2, 3, ...),
    plus which source column names map into each level. Accepted shapes:

      Ordered list of (canonical, [source_aliases]) -- order IS the level:
        [("DEPT",    ["Row labels", "Department"]),   # level 1
         ("Service", ["Division"]),                    # level 2
         ("Variant", ["Type", "Kind"])]                # level 3

      Dict keyed by 1-based level number:
        {1: ("DEPT", ["Row labels"]),
         2: ("Service", ["Division"]),
         3: ("Variant", ["Type"])}

    A canonical name with no aliases is fine (sources already use that name).
    Returns (aliases, hierarchy_levels) ready for parse()/ingest().
    """
    if isinstance(spec, dict):
        items = [spec[k] for k in sorted(spec)]  # order by level number
    else:
        items = list(spec)

    aliases, levels = {}, []
    for entry in items:
        if isinstance(entry, (list, tuple)):
            canonical = entry[0]
            names = list(entry[1]) if len(entry) > 1 and entry[1] else []
        else:
            canonical, names = entry, []
        levels.append(canonical)
        if names:
            aliases[canonical] = names
    return aliases, levels


# SME-facing audit: for each file, shows which SOURCE column landed in each
# business level (and flags any level left empty) -- the sheet you take back to
# a subject-matter expert to confirm the mapping is right.
def hierarchy_audit(report, hierarchy_levels):
    """Summarize, per source file, which source header filled each business
    level. Returns a tidy DataFrame: source_file, business_level, canonical,
    source_header, status ('mapped' | 'MISSING')."""
    has_file = "source_file" in report.columns
    files = report["source_file"].unique() if has_file else [None]
    rows = []
    for f in files:
        sub = report[report["source_file"] == f] if has_file else report
        for i, canonical in enumerate(hierarchy_levels, start=1):
            hits = sub[sub["resolved_name"] == canonical]["source_header"].dropna()
            src = hits.iloc[0] if len(hits) else None
            rows.append({
                "source_file": f,
                "business_level": i,
                "canonical": canonical,
                "source_header": src,
                "status": "mapped" if src is not None else "MISSING",
            })
    cols = ["source_file", "business_level", "canonical", "source_header", "status"]
    out = pd.DataFrame(rows, columns=cols)
    return out if has_file else out.drop(columns=["source_file"])


# Applies the DECLARED business hierarchy: puts the canonical level columns in
# business order and enforces that each file exposes exactly those levels after
# aliasing, so cross-file misalignment surfaces instead of silently passing.
# Every declared level MUST be present as a column (schema is enforced either
# way): appends a HIGH row for any missing or unexpected level. A level absent
# from the source fails HIGH by default; if that exact level is explicitly
# permitted for this report, the column is still created, filled with the
# OVERRIDE_NULL sentinel, marked per-row, and logged as OVERRIDDEN (not silenced).
def _apply_hierarchy(df, hierarchy_levels, var_name, value_name, report=None,
                     allow_missing=(), report_name=None):
    """Order dimension columns to match hierarchy_levels and enforce the schema.

    allow_missing : iterable of canonical level names that this report is
        explicitly permitted to lack. Any level in this list that is absent gets
        created and filled with OVERRIDE_NULL instead of failing.

    Returns (df, report). The output ALWAYS carries every declared level column
    plus NULLED_LEVELS_COLUMN, so the schema is identical whether or not an
    override fired.
    """
    meta = set(HIERARCHY_COLUMNS + [var_name, value_name, "source_file",
                                    NULLED_LEVELS_COLUMN])
    # Metadata provenance columns are namespaced and are NOT dimension columns.
    dims = [c for c in df.columns
            if c not in meta and not str(c).startswith(METADATA_PREFIX)]

    allow = set(allow_missing or ())
    missing = [h for h in hierarchy_levels if h not in df.columns]
    extra = [d for d in dims if d not in hierarchy_levels]

    overridden = [h for h in missing if h in allow]      # permitted -> null it
    unpermitted = [h for h in missing if h not in allow]  # -> HIGH, fail

    # Schema enforcement: create every missing level column regardless. The
    # permitted ones get the sentinel; the unpermitted ones get it too, but they
    # ALSO raise a HIGH flag so the load can be failed on review. This keeps the
    # output shape stable even for a file that is going to be rejected.
    for h in overridden + unpermitted:
        df[h] = OVERRIDE_NULL if h in allow else pd.NA

    # Row-level marker: which levels are nulled by override on this row.
    df[NULLED_LEVELS_COLUMN] = ", ".join(sorted(overridden)) if overridden else ""

    if report is not None and (unpermitted or extra or overridden):
        rows = []
        for h in unpermitted:
            rows.append({
                "source_header": None, "resolved_name": h,
                "method": "hierarchy_missing", "matched_against": None,
                "confidence": None, "review_priority": REVIEW_HIGH,
                "note": f"declared hierarchy level '{h}' not present after "
                        f"aliasing and not permitted-missing for this report",
            })
        for h in overridden:
            rows.append({
                "source_header": None, "resolved_name": h,
                "method": "hierarchy_overridden", "matched_against": None,
                "confidence": None, "review_priority": REVIEW_OVERRIDDEN,
                "note": f"level '{h}' absent but explicitly permitted for report "
                        f"{report_name!r}; column created and filled "
                        f"with {OVERRIDE_NULL!r}",
            })
        for d in extra:
            rows.append({
                "source_header": d, "resolved_name": d,
                "method": "hierarchy_unexpected", "matched_against": None,
                "confidence": None, "review_priority": REVIEW_HIGH,
                "note": f"column '{d}' is not a declared hierarchy level",
            })
        add = pd.DataFrame(rows)
        if "source_file" in report.columns and "source_file" not in add.columns:
            add.insert(0, "source_file", report["source_file"].iloc[0]
                       if len(report) else None)
        report = pd.concat([report, add], ignore_index=True)

    # Reorder into the enforced schema: source_file (if any), ALL declared levels
    # in business order, any extra dims, the nulled-levels marker, the hierarchy
    # metadata, then period + value.
    present_levels = [h for h in hierarchy_levels if h in df.columns]
    meta_cols = [c for c in df.columns if str(c).startswith(METADATA_PREFIX)]
    lead = (["source_file"] if "source_file" in df.columns else []) \
        + meta_cols + present_levels + extra
    ordered = lead + [NULLED_LEVELS_COLUMN] \
        + [c for c in HIERARCHY_COLUMNS if c in df.columns] \
        + [var_name, value_name]
    df = df[[c for c in ordered if c in df.columns]]
    return df, report


# ============================================================================
# Reconciliation — THE ARITHMETIC GATE
# ============================================================================
#
# Mapping errors are loud and cosmetic. Arithmetic errors are silent and
# expensive: a subtly mis-parsed sheet that still "succeeds" flows straight into
# a decision-maker's report. This is the check that makes silent-wrong loud.
#
# The pivot's own `total` row is the source's self-declared answer. We extract it
# BEFORE it is discarded, then assert that our leaf rows sum to it, per period.
# Because leaves and subtotal parents both live in the output, we sum ONLY the
# leaves -- summing everything would double-count and (wrongly) tie to nothing.


# Pulls the pivot's own total row out of the sheet, per period, before the
# parser throws it away. This is the "expected" side of the reconciliation.
def read_total_row(path, sheet=0, aliases=None, fuzzy=True,
                   accept_threshold=DEFAULT_ACCEPT_THRESHOLD,
                   review_threshold=DEFAULT_REVIEW_THRESHOLD,
                   strip_prefixes=DEFAULT_AGG_PREFIXES,
                   total_re=None, value_threshold=DEFAULT_VALUE_THRESHOLD):
    """Return {period_name: total_value} from the sheet's total row.

    Returns {} if the sheet has no total row -- which is itself reportable,
    because without one the arithmetic gate cannot run.
    """
    values, _ind = _read_sheet_with_indent(path, sheet)
    hdr = _find_header_row(values)
    headers = values.iloc[hdr].tolist()
    body = values.iloc[hdr + 1:].reset_index(drop=True)
    body.columns = range(body.shape[1])

    valid_cols, raw_names, records, col_names = _resolve_columns(
        headers, aliases, fuzzy, accept_threshold, review_threshold,
        strip_prefixes)
    body = body[valid_cols]
    body.columns = col_names

    # Anchor the total test to the outline column, exactly as the parsers do.
    # Otherwise a stray "total" in the trailing notes block below the table could
    # become the reconciliation BASELINE -- i.e. the gate would be measuring
    # against the wrong number, which is worse than having no gate at all. This
    # reuses the parsers' own classifier so the baseline is read on the same
    # column split the data was.
    total_re = total_re or TOTAL_RE
    value_cols, dim_cols = _split_value_dims(body, col_names, total_re,
                                             value_threshold)
    outline = dim_cols[0] if dim_cols else None

    total_rows = body[body.apply(
        _is_total_row, axis=1, dim_cols=dim_cols, outline=outline,
        total_re=total_re)]
    if total_rows.empty:
        return {}
    # If several match (subtotal + grand total), the LAST one is the grand total.
    tr = total_rows.iloc[-1]

    totals = {}
    for c in col_names:
        num = to_number(tr[c])
        if pd.notna(num):
            totals[c] = float(num)
    return totals


# THE GATE: asserts the parsed leaf rows sum to the pivot's own total row,
# per period. A mismatch is HIGH and blocks the load -- this is the check that
# catches "parsed successfully, but the numbers are wrong".
def total_validation(data, totals, var_name="Period", value_name="Value",
                     tolerance=RECON_TOLERANCE, source=None):
    """Compare leaf-row sums against the source's declared total row.

    data    : tidy output for ONE source (must carry `is_leaf`).
    totals  : {period: expected_total} from read_total_row().
    Returns (ok, records) -- records are report rows for every mismatch, plus
    one row if the source declared no total at all.

    Only is_leaf rows are summed: parent rows are subtotals of their children,
    so including them would double-count.
    """
    recs = []

    def rec(method, note, resolved=None, priority=REVIEW_HIGH):
        return {
            "source_header": source, "resolved_name": resolved,
            "method": method, "matched_against": None, "confidence": None,
            "review_priority": priority, "note": note,
        }

    if not totals:
        recs.append(rec(
            "reconciliation_no_total",
            "source has no total row; the arithmetic gate cannot verify this "
            "file. Add a total row or explicitly waive reconciliation.",
        ))
        return False, recs

    if "is_leaf" in data.columns:
        leaves = data[data["is_leaf"]]
    else:  # harmonize() guarantees is_leaf, but stay defensive
        leaves = data

    actual = leaves.groupby(var_name)[value_name].sum().to_dict()

    # Every period the source declared a total for must tie.
    for period, expected in totals.items():
        got = actual.get(period)
        if got is None:
            recs.append(rec(
                RECON_METHOD,
                f"period {period!r} has a total of {expected:,.2f} in the source "
                f"but no parsed rows",
                resolved=period,
            ))
            continue
        diff = abs(got - expected)
        if diff > tolerance:
            recs.append(rec(
                RECON_METHOD,
                f"period {period!r}: leaf sum {got:,.2f} != source total "
                f"{expected:,.2f} (diff {diff:,.2f}, tolerance {tolerance})",
                resolved=period,
            ))

    # A period we parsed but the source never totalled is also suspicious.
    for period in actual:
        if period not in totals:
            recs.append(rec(
                RECON_METHOD,
                f"period {period!r} was parsed but has no total in the source "
                f"row; column may be misaligned",
                resolved=period,
            ))

    return (len(recs) == 0), recs


# Single front door: figures out which parse class the file needs and runs it.
def parse(
    path: str,
    sheet=None,
    value_name: str = None,
    var_name: str = None,
    aliases=None,
    fuzzy: bool = True,
    accept_threshold: float = DEFAULT_ACCEPT_THRESHOLD,
    review_threshold: float = DEFAULT_REVIEW_THRESHOLD,
    strip_prefixes=None,
    return_report: bool = False,
    mode: str = "auto",
    hierarchy_col=None,
    harmonize: bool = True,
    hierarchy_levels=None,
    hierarchy=None,
    allow_missing_levels=(),
    check_report_name: bool = True,
    metadata_fields=DEFAULT_METADATA_FIELDS,
    required_metadata=REQUIRED_METADATA_FIELDS,
    validate_totals: bool = True,
    recon_tolerance: float = RECON_TOLERANCE,
    config=None,
    vocabularies=None,
    licenses=None,
    total_re=None,
    value_threshold=None,
    blank_literal=None,
    metadata_scan_rows=None,
    modes=None,
    registry=None,
    bootstrap=None,
    parents=None,
    drift_tolerance=None,
):
    """Auto-routing entry point. Detects whether the sheet is an indented
    (hierarchical) pivot and dispatches to the right parser.

    config : the SME declaration -- a path to configuration.json or an already-
        loaded dict from load_configuration(). Supplies aliases, ordered levels,
        member vocabularies and normalization licenses in one artifact. Anything
        passed explicitly alongside it wins, so a caller can still override.

    mode : 'auto' (default) sniffs the sheet; 'flat' or 'hierarchical' force it.
    harmonize : when True (default), every result carries the same columns in
        the same order (HIERARCHY_COLUMNS included), so flat and hierarchical
        outputs share one schema and stack cleanly across a mixed batch.
    hierarchy : a single upfront business-hierarchy spec (see build_hierarchy).
        When given, it supplies BOTH the ordered level names and the source
        aliases at once -- the one artifact you maintain with SMEs.
    hierarchy_levels : ordered list of CANONICAL level names. Usually supplied
        via `hierarchy` instead.
    allow_missing_levels : levels this report is EXPLICITLY permitted to lack.
        Default () -- a missing level is a HIGH flag and the load should fail.
        A level named here is instead created, filled with OVERRIDE_NULL
        ("NA - Overridden"), marked per-row in `nulled_levels`, and logged as
        OVERRIDDEN. The output SCHEMA is identical either way.
    check_report_name : enforce the input contract that the 'Report:' cell above
        the pivot exists and exactly equals the sheet name. Violations are HIGH.

    Signature is ingest-compatible, so you can pass parse itself as ingest's
    `parser` to auto-route a whole batch.
    """
    # The DECLARATION, if given, supplies aliases + ordered levels + member
    # vocabularies + normalization licenses as ONE artifact. Explicit arguments
    # still win over it, so a caller can override a single field for one run.
    cfg = _expand_config(config)
    if cfg:
        aliases = {**cfg["aliases"], **(aliases or {})}
        hierarchy_levels = hierarchy_levels or cfg["hierarchy_levels"]
        vocabularies = vocabularies or cfg["vocabularies"]
        licenses = licenses or cfg["licenses"]
        modes = modes or cfg["modes"]
        bootstrap = bootstrap or cfg["bootstrap"]
        parents = parents or cfg["parents"]
        drift_tolerance = drift_tolerance or cfg["drift_tolerance"]
        # A registry level needs the register. Load it lazily -- a config with no
        # registry levels never touches the file.
        if registry is None and MODE_REGISTRY in set(cfg["modes"].values()):
            registry = load_registry(cfg["registry_path"])
        # format semantics + output shape: the call site carries none of this
        strip_prefixes = strip_prefixes or cfg["strip_prefixes"]
        total_re = total_re or cfg["total_re"]
        value_threshold = value_threshold or cfg["value_threshold"]
        blank_literal = blank_literal or cfg["blank_literal"]
        metadata_scan_rows = metadata_scan_rows or cfg["metadata_scan_rows"]
        var_name = var_name or cfg["var_name"]
        value_name = value_name or cfg["value_name"]
        sheet = cfg["sheet"] if sheet is None else sheet

    # Fall back to the module defaults for anything still unset.
    sheet = 0 if sheet is None else sheet
    var_name = var_name or "Period"
    value_name = value_name or "Value"
    strip_prefixes = strip_prefixes or DEFAULT_AGG_PREFIXES
    total_re = total_re or TOTAL_RE
    value_threshold = value_threshold or DEFAULT_VALUE_THRESHOLD
    blank_literal = blank_literal or BLANK_MEMBER_LITERAL
    metadata_scan_rows = metadata_scan_rows or METADATA_SCAN_ROWS
    drift_tolerance = drift_tolerance or DEFAULT_DRIFT_TOLERANCE

    # A single hierarchy spec expands into aliases + ordered levels.
    if hierarchy is not None:
        spec_aliases, spec_levels = build_hierarchy(hierarchy)
        aliases = {**spec_aliases, **(aliases or {})}
        hierarchy_levels = hierarchy_levels or spec_levels
    if mode == "auto":
        mode = "hierarchical" if _sheet_is_hierarchical(path, sheet) else "flat"

    if mode == "hierarchical":
        result = extract_hierarchical_pivot(
            path, sheet=sheet, value_name=value_name, var_name=var_name,
            hierarchy_col=hierarchy_col, aliases=aliases, fuzzy=fuzzy,
            accept_threshold=accept_threshold, review_threshold=review_threshold,
            strip_prefixes=strip_prefixes, return_report=return_report,
            vocabularies=vocabularies, licenses=licenses,
            total_re=total_re, value_threshold=value_threshold,
            blank_literal=blank_literal, modes=modes, registry=registry,
            bootstrap=bootstrap, parents=parents,
            drift_tolerance=drift_tolerance,
        )
    else:
        result = extract_pivot(
            path, sheet=sheet, value_name=value_name, var_name=var_name,
            aliases=aliases, fuzzy=fuzzy, accept_threshold=accept_threshold,
            review_threshold=review_threshold, strip_prefixes=strip_prefixes,
            return_report=return_report,
            vocabularies=vocabularies, licenses=licenses,
            total_re=total_re, value_threshold=value_threshold,
            blank_literal=blank_literal, modes=modes, registry=registry,
            bootstrap=bootstrap, parents=parents,
            drift_tolerance=drift_tolerance,
        )

    if not harmonize:
        return result

    # --- Metadata contract: read the whole block, validate, carry as provenance.
    name_recs, report_name, meta_values = [], None, {}
    if check_report_name:
        try:
            values, _ind = _read_sheet_with_indent(path, sheet)
            hdr = _find_header_row(values)
            block, _anchor = read_metadata_block(
                values, hdr, scan_rows=metadata_scan_rows)
            sheet_name = _sheet_name_of(path, sheet)
            name_recs, meta_values = validate_metadata(
                block, sheet_name,
                declared=metadata_fields, required=required_metadata,
            )
            report_name = meta_values.get("Report")
        except Exception as e:  # noqa: BLE001 - contract check must not crash parse
            name_recs = [{
                "source_header": None, "resolved_name": None,
                "method": "metadata_check_failed", "matched_against": None,
                "confidence": None, "review_priority": REVIEW_HIGH,
                "note": f"{type(e).__name__}: {e}",
            }]
            meta_values = {f: pd.NA for f in metadata_fields}

    def _finish(tidy, report):
        if name_recs and report is not None:
            add = pd.DataFrame(name_recs)
            if "source_file" in report.columns and "source_file" not in add.columns:
                add.insert(0, "source_file", report["source_file"].iloc[0]
                           if len(report) else None)
            report = pd.concat([report, add], ignore_index=True)
        tidy = _harmonize(tidy, var_name, value_name)
        # Attach metadata as provenance columns. Every DECLARED field always
        # becomes a column (pd.NA when absent), so the schema never drifts.
        if check_report_name:
            for field in metadata_fields:
                tidy[_meta_column(field)] = meta_values.get(field, pd.NA)
        if hierarchy_levels:
            tidy, report = _apply_hierarchy(
                tidy, hierarchy_levels, var_name, value_name, report,
                allow_missing=allow_missing_levels, report_name=report_name,
            )

            # THE DIMENSION GATE: every row must be addressable at every declared
            # level. Runs AFTER the repeat-fill, so it judges what actually comes
            # out. Independent of WHY a key is null -- suppressed repeat we could
            # not resolve, bad alias, hole upstream -- because a null key silently
            # drops out of every GROUP BY downstream and nobody sees it go.
            # OVERRIDE_NULL is a declared value, not a null, so overrides pass.
            if report is not None:
                dim_recs = validate_dimension_completeness(tidy, hierarchy_levels)
                if dim_recs:
                    add = pd.DataFrame(dim_recs, columns=REPORT_COLUMNS)
                    if "source_file" in report.columns:
                        add.insert(0, "source_file", report["source_file"].iloc[0]
                                   if len(report) else None)
                    report = pd.concat([report, add], ignore_index=True)

        # THE ARITHMETIC GATE: leaf sums must tie to the source's own total row.
        # Runs last, on the final parsed frame, so it validates what actually
        # comes out -- not an intermediate. A mismatch is HIGH and blocks the load.
        if validate_totals and report is not None:
            try:
                totals = read_total_row(
                    path, sheet=sheet, aliases=aliases, fuzzy=fuzzy,
                    accept_threshold=accept_threshold,
                    review_threshold=review_threshold,
                    strip_prefixes=strip_prefixes,
                    total_re=total_re, value_threshold=value_threshold,
                )
                _ok, recon_recs = total_validation(
                    tidy, totals, var_name=var_name, value_name=value_name,
                    tolerance=recon_tolerance, source=report_name,
                )
            except Exception as e:  # noqa: BLE001 - a failed gate must not pass silently
                recon_recs = [{
                    "source_header": None, "resolved_name": None,
                    "method": "reconciliation_failed", "matched_against": None,
                    "confidence": None, "review_priority": REVIEW_HIGH,
                    "note": f"could not verify totals: {type(e).__name__}: {e}",
                }]
            if recon_recs:
                add = pd.DataFrame(recon_recs)
                if "source_file" in report.columns \
                        and "source_file" not in add.columns:
                    add.insert(0, "source_file",
                               report["source_file"].iloc[0] if len(report) else None)
                report = pd.concat([report, add], ignore_index=True)

        return tidy, report

    if return_report:
        tidy, report = result
        return _finish(tidy, report)

    tidy, _ = _finish(result, None)
    return tidy


if __name__ == "__main__":
    # Demo: define the business hierarchy ONCE (the artifact you build with
    # SMEs), then mass-ingest a mixed batch. Order = business level 1 > 2 > 3;
    # each level lists the source column names that map into it. Different
    # sources call level 2 "Service" or "Division" -- both land in Service.
    BUSINESS_HIERARCHY = [
        ("DEPT",    ["Row labels", "Department"]),  # level 1
        ("Service", ["Division"]),                   # level 2
        ("Variant", ["Type"]),                       # level 3
    ]
    files = [
        "/mnt/user-data/uploads/Extraction_Sample_1.xlsx",  # flat, native names
        "/mnt/user-data/uploads/Extraction_Sample_2.xlsx",  # flat, pivot names
        "/mnt/user-data/uploads/Extraction_Sample_3.xlsx",  # indented sub-levels
        "/mnt/user-data/uploads/does_not_exist.xlsx",       # -> file_error row
    ]
    data, report = ingest(
        files, hierarchy=BUSINESS_HIERARCHY,
        var_name="FiscalYear", value_name="Value",
    )

    pd.set_option("display.max_rows", 100)
    pd.set_option("display.width", 180)
    print(f"Ingested rows: {len(data)}  |  files: {data['source_file'].nunique()}")
    print(f"Columns: {list(data.columns)}")

    print("\n=== SME HIERARCHY AUDIT (which source column -> which level) ===")
    print(hierarchy_audit(report, ["DEPT", "Service", "Variant"]).to_string(index=False))

    high = review_items(report)
    print(f"\n=== HIGH-PRIORITY REVIEW ({len(high)}) ===")
    print(high.to_string(index=False) if len(high) else "  none")