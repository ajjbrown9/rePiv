"""
make_fixtures.py
----------------
Regenerate the 10 deliberate mutation fixtures the regression suite feeds through
pivot_extract. Each is derived from Extraction_Sample_6.xlsx so the "clean"
baseline (metadata block, header, notes) is identical and only the one property
under test is mutated.

Layout produced (relative to CWD, matching test_pivot_extract.py):
  m1_real_excel_indent.xlsx      m2_blanked_repeat_labels.xlsx
  m3_paren_negatives.xlsx        m4_dept_named_Total.xlsx
  m5_numbers_as_text.xlsx
  mut/M1_compact_form.xlsx       mut/M2_numeric_dims.xlsx
  mut/M3_subtotal_rows.xlsx      mut/M4_deep_nest_blank.xlsx
  mut/M5_dirty_dim_values.xlsx
"""
import os
import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

SRC = "Extraction_Sample_6.xlsx"
SHEET = "F-DEPT-VAR-TEAM-SPRL"
os.makedirs("mut", exist_ok=True)

# ---- capture the original grid (values + indent) ---------------------------
wb0 = load_workbook(SRC, data_only=True)
ws0 = wb0[SHEET]
MAXR, MAXC = ws0.max_row, ws0.max_column
VAL = {(c.row, c.column): c.value for row in ws0.iter_rows() for c in row}
IND = {(c.row, c.column): int(c.alignment.indent or 0)
       for row in ws0.iter_rows() for c in row}

HEADER_ROW = 12
FIRST_DATA, LAST_DATA = 13, 32
TOTAL_ROW = 33
VALUE_COLS = list(range(4, 11))   # D..J  -> FY26..FY32 (7 periods)
PERIODS = [VAL[(HEADER_ROW, c)] for c in VALUE_COLS]

# Business grouping of the 20 data rows by top-level DEPT (row indices).
A_PARENT, A_SUB1, A_SUB2 = 13, 14, 15
A_FLAT = [16, 17, 18, 19, 20]          # A rows that are their own leaves
B_ROWS = [21, 22, 23, 24, 25, 26]
C_ROWS = [27, 28, 29, 30, 31, 32]


def clone(dst):
    """Byte-copy the original so unmutated structure/formatting is preserved."""
    shutil.copy(SRC, dst)
    return load_workbook(dst)


def scaffold():
    """Fresh workbook carrying the original metadata/junk rows 1-11 (cols A-C).
    Header + data are written by the caller."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    for r in range(1, HEADER_ROW):                 # rows 1..11
        for c in range(1, 4):                      # cols A,B,C
            v = VAL.get((r, c))
            if v is not None:
                ws.cell(row=r, column=c, value=v)
    return wb, ws


def notes(ws, r):
    """Append the trailing NOTES/Comments rows (text only, no numbers)."""
    ws.cell(row=r, column=1,
            value="*NOTES regenerated fixture; text-only trailing block")
    ws.cell(row=r + 1, column=1, value="Comments: fixture")


# ===========================================================================
# ROUND 1
# ===========================================================================

def m1_real_excel_indent():
    """Subtypes indented via the NATIVE Excel alignment.indent property instead
    of leading spaces. Must parse clean and reconcile."""
    wb = clone("m1_real_excel_indent.xlsx")
    ws = wb[SHEET]
    for r in (A_SUB1, A_SUB2):
        cell = ws.cell(row=r, column=1)
        cell.value = str(VAL[(r, 1)]).strip()      # drop the leading spaces
        cell.alignment = Alignment(indent=1)       # real outline level instead
    wb.save("m1_real_excel_indent.xlsx")


def m2_blanked_repeat_labels():
    """Excel default 'Repeat All Item Labels' OFF: a repeated top-level DEPT
    prints only on its first row and is blank thereafter. The dangerous case is
    a blank level-0 label sitting directly under an indented child."""
    wb = clone("m2_blanked_repeat_labels.xlsx")
    ws = wb[SHEET]
    last_top = None
    for r in range(FIRST_DATA, LAST_DATA + 1):
        raw = VAL[(r, 1)]
        is_sub = isinstance(raw, str) and raw != raw.lstrip(" ")   # indented child
        if is_sub:
            continue                                # unique subtype labels stay
        if raw == last_top:
            ws.cell(row=r, column=1).value = None   # suppress repeated label
        else:
            last_top = raw
    wb.save("m2_blanked_repeat_labels.xlsx")


def m3_paren_negatives():
    """One value written as an accounting negative '(948)'. Total row adjusted so
    the file reconciles to its own new baseline (86273 - 2*948 = 84377)."""
    wb = clone("m3_paren_negatives.xlsx")
    ws = wb[SHEET]
    r = 17                                          # A / F / C, a leaf, FY26=948
    ws.cell(row=r, column=4, value="(948)")         # -> -948
    new_total = VAL[(TOTAL_ROW, 4)] - 2 * 948       # 86273 -> 84377
    ws.cell(row=TOTAL_ROW, column=4, value=new_total)
    wb.save("m3_paren_negatives.xlsx")


def m4_dept_named_Total():
    """A real department literally named 'Total Rewards' (contains the word
    'total'). Must NOT be deleted by total-row detection; blocks only because
    the member is undeclared."""
    wb = clone("m4_dept_named_Total.xlsx")
    ws = wb[SHEET]
    for r in B_ROWS:                                # 6 rows, all keep Service/Variant
        ws.cell(row=r, column=1, value="Total Rewards")
    wb.save("m4_dept_named_Total.xlsx")


def m5_numbers_as_text():
    """Every value stored as text with thousands separators ('2,822',
    '4,359.5'). Parser must coerce and reconcile."""
    wb = clone("m5_numbers_as_text.xlsx")
    ws = wb[SHEET]
    for r in list(range(FIRST_DATA, LAST_DATA + 1)) + [TOTAL_ROW]:
        for c in VALUE_COLS:
            v = VAL[(r, c)]
            if isinstance(v, (int, float)):
                ws.cell(row=r, column=c, value=f"{v:,}")
    wb.save("m5_numbers_as_text.xlsx")


# ===========================================================================
# ROUND 2
# ===========================================================================

def M1_compact_form():
    """All category columns collapsed away: only the indented Row-labels column
    and the value columns remain. Service and Variant do not exist as columns."""
    wb, ws = scaffold()
    # header: Row labels + the 7 value headers, shifted to start at col B
    ws.cell(row=HEADER_ROW, column=1, value=VAL[(HEADER_ROW, 1)])   # 'Row labels'
    for j, c in enumerate(VALUE_COLS):
        ws.cell(row=HEADER_ROW, column=2 + j, value=VAL[(HEADER_ROW, c)])
    # data: outline label (indent preserved via original leading spaces) + values
    out_r = FIRST_DATA
    for r in range(FIRST_DATA, LAST_DATA + 1):
        ws.cell(row=out_r, column=1, value=VAL[(r, 1)])
        for j, c in enumerate(VALUE_COLS):
            ws.cell(row=out_r, column=2 + j, value=VAL[(r, c)])
        out_r += 1
    # total row
    ws.cell(row=out_r, column=1, value=VAL[(TOTAL_ROW, 1)])
    for j, c in enumerate(VALUE_COLS):
        ws.cell(row=out_r, column=2 + j, value=VAL[(TOTAL_ROW, c)])
    notes(ws, out_r + 2)
    wb.save("mut/M1_compact_form.xlsx")


def M2_numeric_dims():
    """Service (Division) coded numerically (10,20,30...). The column looks like
    data, gets melted, and the Service dimension vanishes."""
    wb = clone("mut/M2_numeric_dims.xlsx")
    ws = wb[SHEET]
    code = {"MC": 10, "N": 20, "F": 30, "F1": 31, "F2": 32, "F3": 33}
    for r in range(FIRST_DATA, LAST_DATA + 1):
        div = VAL[(r, 2)]
        if div in code:
            ws.cell(row=r, column=2, value=code[div])
    wb.save("mut/M2_numeric_dims.xlsx")


def M3_subtotal_rows():
    """Flat 18-row table (no subtypes) with per-DEPT Subtotal rows plus a final
    Grand Total. Subtotals dropped; the LAST total row is the baseline."""
    wb, ws = scaffold()
    for c in range(1, 4):
        ws.cell(row=HEADER_ROW, column=c, value=VAL[(HEADER_ROW, c)])
    for j, c in enumerate(VALUE_COLS):
        ws.cell(row=HEADER_ROW, column=4 + j, value=VAL[(HEADER_ROW, c)])

    def vals(r):
        return [VAL[(r, c)] for c in VALUE_COLS]

    def put(row, dept, service, variant, values):
        ws.cell(row=row, column=1, value=dept)
        ws.cell(row=row, column=2, value=service)
        ws.cell(row=row, column=3, value=variant)
        for j, v in enumerate(values):
            ws.cell(row=row, column=4 + j, value=v)

    # A collapsed to 6 flat rows: use the PARENT value for A/MC/B (== sum of its
    # two subtypes), then A's five other rows verbatim.
    A_rows = [(VAL[(A_PARENT, 2)], VAL[(A_PARENT, 3)], vals(A_PARENT))] + \
             [(VAL[(r, 2)], VAL[(r, 3)], vals(r)) for r in A_FLAT]
    B_rows_ = [(VAL[(r, 2)], VAL[(r, 3)], vals(r)) for r in B_ROWS]
    C_rows_ = [(VAL[(r, 2)], VAL[(r, 3)], vals(r)) for r in C_ROWS]

    row = HEADER_ROW + 1
    for dept, group in (("A", A_rows), ("B", B_rows_), ("C", C_rows_)):
        for svc, var, vv in group:
            put(row, dept, svc, var, vv)
            row += 1
        subtotal = [sum(g[2][k] for g in group) for k in range(len(VALUE_COLS))]
        put(row, "Subtotal", None, None, subtotal)   # dropped by the parser
        row += 1
    grand = [VAL[(TOTAL_ROW, c)] for c in VALUE_COLS]
    put(row, "Grand Total", None, None, grand)        # LAST total = baseline
    notes(ws, row + 2)
    wb.save("mut/M3_subtotal_rows.xlsx")


def M4_deep_nest_blank():
    """Three indent levels: A > A subtype 1 > {1a, 1b}. The sub-sub members are
    undeclared (a new member is a business decision). One Service cell holds the
    literal '(blank)' -> mapped to the sentinel."""
    wb = clone("mut/M4_deep_nest_blank.xlsx")
    ws = wb[SHEET]
    # Insert two level-2 rows immediately after A subtype 1 (row 14).
    ws.insert_rows(A_SUB1 + 1, amount=2)
    for k, name in enumerate(("A subtype 1a", "A subtype 1b")):
        r = A_SUB1 + 1 + k
        ws.cell(row=r, column=1, value="              " + name)   # 14 spaces = lvl2
        ws.cell(row=r, column=2, value=VAL[(A_SUB1, 2)])          # MC
        ws.cell(row=r, column=3, value=VAL[(A_SUB1, 3)])          # B
        for c in VALUE_COLS:
            ws.cell(row=r, column=c, value=VAL[(A_SUB1, c)] / 2)  # even split
    # A subtype 2 has shifted down by 2 (was row 15 -> now 17); blank its Service.
    ws.cell(row=A_SUB2 + 2, column=2, value="(blank)")
    wb.save("mut/M4_deep_nest_blank.xlsx")


def M5_dirty_dim_values():
    """THE HEADLINE. Three departments in, six out. Whitespace/unicode variants
    heal silently ('A ' -> A, 'C\\xa0' -> C); a case variant ('b') is a decision,
    stays undeclared, and blocks -- while the arithmetic ties exactly.

    Built FLAT (no indent). The hierarchical parser strips the outline column
    during indent extraction, which would silently clean the whitespace variants
    before the identity gate ever sees them; on a flat sheet the dirty key
    survives to validate_members, which is the whole point of the fixture."""
    wb, ws = scaffold()
    for c in range(1, 4):
        ws.cell(row=HEADER_ROW, column=c, value=VAL[(HEADER_ROW, c)])
    for j, c in enumerate(VALUE_COLS):
        ws.cell(row=HEADER_ROW, column=4 + j, value=VAL[(HEADER_ROW, c)])

    def vals(r):
        return [VAL[(r, c)] for c in VALUE_COLS]

    # 18 flat rows: A collapsed to its parent value for A/MC/B, then the rest.
    A_src = [A_PARENT] + A_FLAT
    rows = ([("A", r) for r in A_src]
            + [("B", r) for r in B_ROWS]
            + [("C", r) for r in C_ROWS])

    # Dirty the DEPT key: two heal-silently variants + one undeclared case variant.
    dept_override = {
        0: "A ",           # A/MC/B  -> trailing space, heals to 'A'
        6: "b", 7: "b", 8: "b",   # three B rows -> lowercase, undeclared
        12: "C\u00a0",     # C/MC/B  -> nbsp, heals to 'C'
    }
    row = HEADER_ROW + 1
    for i, (dept, r) in enumerate(rows):
        ws.cell(row=row, column=1, value=dept_override.get(i, dept))
        ws.cell(row=row, column=2, value=VAL[(r, 2)])
        ws.cell(row=row, column=3, value=VAL[(r, 3)])
        for j, v in enumerate(vals(r)):
            ws.cell(row=row, column=4 + j, value=v)
        row += 1
    # single grand total row (baseline)
    ws.cell(row=row, column=1, value="total")
    for j, c in enumerate(VALUE_COLS):
        ws.cell(row=row, column=4 + j, value=VAL[(TOTAL_ROW, c)])
    notes(ws, row + 2)
    wb.save("mut/M5_dirty_dim_values.xlsx")


if __name__ == "__main__":
    for fn in (m1_real_excel_indent, m2_blanked_repeat_labels, m3_paren_negatives,
               m4_dept_named_Total, m5_numbers_as_text,
               M1_compact_form, M2_numeric_dims, M3_subtotal_rows,
               M4_deep_nest_blank, M5_dirty_dim_values):
        fn()
        print("wrote", fn.__name__)
